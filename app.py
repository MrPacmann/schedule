"""Streamlit-приложение для поиска занятий преподавателя в Excel-расписании."""

from __future__ import annotations

import logging
import math
import re
from collections import Counter
from collections.abc import MutableMapping, Sequence
from dataclasses import dataclass
from datetime import datetime, time
from hashlib import sha256
from io import BytesIO
from numbers import Real
from typing import Any, BinaryIO, Final

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LOGGER = logging.getLogger(__name__)

GROUP_PATTERN: Final[re.Pattern[str]] = re.compile(r"[А-Я]{4}-\d{2}-\d{2}")
WHITESPACE_PATTERN: Final[re.Pattern[str]] = re.compile(r"\s+")
PAIR_NUMBER_PATTERN: Final[re.Pattern[str]] = re.compile(r"\d+(?:[.,]\d+)?")
TIME_TEXT_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"^(?P<hours>\d{1,2})\s*[:.\-]\s*(?P<minutes>\d{2})"
    r"(?:\s*[:.\-]\s*(?P<seconds>\d{2}))?$"
)
WORKLOAD_SUBJECT_SUFFIX_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"\s*\((?=[^)]*(?:з\.?\s*е\.?|\bч\.?|УП\s*№|Д\s*№))[^)]*\)\s*$",
    re.IGNORECASE,
)
NAME_TOKEN_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"[а-яёa-z0-9+#]+",
    re.IGNORECASE,
)
WEEK_SEQUENCE_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"(?<!\d)(?P<weeks>\d{1,2}(?:\s*(?:[,;]|[-–—])\s*\d{1,2})*)"
    r"\s*н(?:\.|\b)",
    re.IGNORECASE,
)

HEADER_SEARCH_ROWS: Final[int] = 10
COMMON_COLUMNS_COUNT: Final[int] = 5
FIRST_GROUP_COLUMN: Final[int] = 5
GROUP_BLOCK_SIZE: Final[int] = 5
SESSION_CACHE_KEY: Final[str] = "_schedule_excel_cache"
PARSED_SCHEDULE_CACHE_KEY: Final[str] = "_parsed_schedule_cache"
MAX_CACHED_FILES: Final[int] = 20
APP_VERSION: Final[str] = "1.8.4"
APP_CHANGELOG: Final[tuple[tuple[str, str, tuple[str, ...]], ...]] = (
    (
        "1.8.4",
        "29.08.2026",
        (
            "Замечания нагрузки объединяются по исходным строкам-потокам.",
            "Группы одного потока показываются в одной строке.",
            "Счётчики сверки теперь отражают строки исходного файла нагрузки.",
        ),
    ),
    (
        "1.8.3",
        "29.08.2026",
        (
            "Переданные другому преподавателю занятия помечаются в расписании сайта и XLSX.",
            "Колонка статуса переименована в «Проверка нагрузки».",
        ),
    ),
    (
        "1.8.2",
        "29.08.2026",
        (
            "Сверка нагрузки стала двусторонней.",
            "Показываются занятия, оставшиеся в расписании у преподавателя после передачи нагрузки другому.",
        ),
    ),
    (
        "1.8.1",
        "29.08.2026",
        ("В возможных местах расписания теперь указывается аудитория занятия.",),
    ),
    (
        "1.8.0",
        "28.08.2026",
        (
            "Ускорен повторный поиск по уже загруженным расписаниям.",
            "Почти вдвое уменьшена память кэша одной серверной сессии.",
            "Проверка аудиторий ограничена слотами выбранных преподавателей.",
            "Накладки преподавателей учитывают конкретные учебные недели.",
            "Исправлена высота многострочных записей в таблицах сайта.",
            "Текст XLSX защищён от интерпретации как формулы.",
            "Минимальная версия Streamlit синхронизирована с интерфейсом.",
        ),
    ),
    (
        "1.7.0",
        "28.08.2026",
        (
            "Добавлена проверка занятости аудиторий.",
            "Конфликты кабинетов ищутся по всем загруженным расписаниям.",
            "Общая лекция одного предмета для нескольких групп не считается ошибкой.",
            "Учитываются конкретные номера учебных недель внутри дисциплины.",
            "Накладки аудиторий добавляются отдельным листом в XLSX.",
        ),
    ),
    (
        "1.6.0",
        "28.08.2026",
        (
            "Накладки поддерживают три и более одновременных занятий.",
            "Практики разных групп больше не объединяются в одно занятие.",
            "Общие лекции нескольких групп по-прежнему не считаются накладкой.",
            "Возможные места из нагрузки агрегируются по временному слоту.",
        ),
    ),
    (
        "1.5.0",
        "28.08.2026",
        (
            "Сверка нагрузки ограничена только введёнными преподавателями.",
            "Однозначные подстановки ФИО добавляются в XLSX и помечаются жёлтым.",
            "Возможные накладки выделяются в XLSX красным цветом.",
            "Расписание, ошибки, подстановки и накладки можно сворачивать.",
        ),
    ),
    (
        "1.4.0",
        "28.08.2026",
        (
            "Добавлена необязательная сверка с файлом учебной нагрузки.",
            "Добавлен выбор осеннего или весеннего семестра.",
            "Ошибки нагрузки и возможные места для ФИО показаны отдельно.",
        ),
    ),
    (
        "1.3.0",
        "28.08.2026",
        (
            "Добавлено постоянное отображение версии приложения.",
            "Добавлен кликабельный журнал изменений в нижней части страницы.",
        ),
    ),
    (
        "1.2.0",
        "28.08.2026",
        (
            "Общие лекции для нескольких групп объединяются без предупреждения.",
            "Разные занятия одного временного слота выводятся в одной строке.",
            "Накладки выделяются в таблице сайта и XLSX-выгрузке.",
        ),
    ),
    (
        "1.1.0",
        "27.08.2026",
        (
            "Добавлена пакетная загрузка файлов разных курсов.",
            "Добавлен поиск сразу нескольких преподавателей.",
            "Добавлена цветная XLSX-выгрузка расписания.",
        ),
    ),
    (
        "1.0.0",
        "27.08.2026",
        ("Первая версия парсера грязных Excel-файлов расписания.",),
    ),
)

QUERY_COLUMN: Final[str] = "Искомый преподаватель"
TEACHER_COLUMN: Final[str] = "Преподаватель"
SOURCE_COLUMN: Final[str] = "Файл-источник"
CONFLICT_COLUMN: Final[str] = "Накладка"
SUGGESTED_COLUMN: Final[str] = "Проверка нагрузки"
ROOM_CONFLICT_COLUMN: Final[str] = "Накладка аудитории"
WORKLOAD_CLASS_TYPES: Final[frozenset[str]] = frozenset({"ЛК", "ПР"})
REMOTE_ROOM_MARKERS: Final[tuple[str, ...]] = (
    "дистанцион",
    "онлайн",
    "online",
    "сдо",
    "zoom",
)
NON_SPECIFIC_ROOM_MARKERS: Final[tuple[str, ...]] = (
    "кафедра",
    "база",
)
NON_SPECIFIC_ROOM_VALUES: Final[frozenset[str]] = frozenset({"вуц"})
EXCLUDED_ROOM_MARKERS: Final[tuple[str, ...]] = (
    *REMOTE_ROOM_MARKERS,
    *NON_SPECIFIC_ROOM_MARKERS,
)

OUTPUT_COLUMNS: Final[list[str]] = [
    "День недели",
    "Пара",
    "Время",
    "Неделя",
    "Группа",
    "Дисциплина",
    "Вид занятий",
    "Аудитория",
]

INTERNAL_COLUMNS: Final[list[str]] = [
    QUERY_COLUMN,
    TEACHER_COLUMN,
    SUGGESTED_COLUMN,
    *OUTPUT_COLUMNS,
    SOURCE_COLUMN,
]

ALL_SCHEDULE_COLUMNS: Final[list[str]] = [
    TEACHER_COLUMN,
    *OUTPUT_COLUMNS,
    SOURCE_COLUMN,
]

WORKLOAD_COLUMNS: Final[list[str]] = [
    "Преподаватель нагрузки",
    "Дисциплина",
    "Вид занятий",
    "Группа",
    "Семестр",
    "Строка нагрузки",
]
WORKLOAD_ROWS_KEY_COLUMN: Final[str] = "__workload_rows"

WORKLOAD_ISSUE_COLUMNS: Final[list[str]] = [
    "Статус",
    "Проблема",
    "Преподаватели по нагрузке",
    "Дисциплина",
    "Вид занятий",
    "Группа",
    "Семестр",
    "ФИО в расписании",
    "Предлагаемое ФИО",
    "Возможная накладка",
    "Возможное место в расписании",
]

ROOM_CONFLICT_COLUMNS: Final[list[str]] = [
    QUERY_COLUMN,
    "День недели",
    "Пара",
    "Время",
    "Неделя",
    "Учебные недели",
    "Аудитория",
    ROOM_CONFLICT_COLUMN,
    "Преподаватели",
    "Группы",
    "Дисциплины",
    "Виды занятий",
    SOURCE_COLUMN,
]

DISPLAY_COLUMNS: Final[list[str]] = [
    QUERY_COLUMN,
    TEACHER_COLUMN,
    SUGGESTED_COLUMN,
    "День недели",
    "Пара",
    "Время",
    "Неделя",
    CONFLICT_COLUMN,
    "Группа",
    "Дисциплина",
    "Вид занятий",
    "Аудитория",
    SOURCE_COLUMN,
]

EXPORT_SUBHEADERS: Final[tuple[str, str, str, str]] = (
    "Дисциплина",
    "Вид",
    "Группа",
    "Аудитория",
)

DAY_EXPORT_CONFIG: Final[tuple[tuple[str, str, str, str], ...]] = (
    ("ПН", "FFFF54", "FBE6A2", "F9DA78"),
    ("ВТ", "75FB4C", "BCD6AC", "9DC284"),
    ("СР", "5884E1", "AAC1F0", "789DE5"),
    ("ЧТ", "F6B26B", "FCE5CD", "F9CB9C"),
    ("ПТ", "B57EDC", "E4D3F3", "CAB2E4"),
    ("СБ", "A6A6A6", "F2F2F2", "D9D9D9"),
)

TEACHER_QUERY_SEPARATOR_PATTERN: Final[re.Pattern[str]] = re.compile(r"[,;\n]+")
TEACHER_NAME_SEPARATOR_PATTERN: Final[re.Pattern[str]] = re.compile(r"[,;]+")

DAY_ORDER: Final[dict[str, int]] = {
    "понедельник": 0,
    "пн": 0,
    "вторник": 1,
    "вт": 1,
    "среда": 2,
    "ср": 2,
    "четверг": 3,
    "чт": 3,
    "пятница": 4,
    "пт": 4,
    "суббота": 5,
    "сб": 5,
}


class ScheduleError(Exception):
    """Базовая ожидаемая ошибка обработки расписания."""


class ScheduleReadError(ScheduleError):
    """Excel-файл не удалось прочитать."""


class ScheduleFormatError(ScheduleError):
    """Структура Excel-файла не похожа на ожидаемое расписание."""


class ScheduleExportError(ScheduleError):
    """Не удалось сформировать Excel-файл с итоговым расписанием."""


class WorkloadError(ScheduleError):
    """Базовая ожидаемая ошибка файла учебной нагрузки."""


class WorkloadFormatError(WorkloadError):
    """Структура файла нагрузки не содержит обязательных столбцов."""


@dataclass(frozen=True, slots=True)
class BatchFileError:
    """Ошибка одного файла внутри пакетной обработки."""

    filename: str
    message: str


@dataclass(frozen=True, slots=True)
class BatchParseResult:
    """Результат отказоустойчивой обработки набора Excel-файлов."""

    records: pd.DataFrame
    all_records: pd.DataFrame
    processed_files: tuple[str, ...]
    duplicate_files: tuple[str, ...]
    errors: tuple[BatchFileError, ...]


@dataclass(frozen=True, slots=True)
class WorkloadAuditResult:
    """Результат сверки учебной нагрузки с расписанием."""

    checked_assignments: int
    matched_assignments: int
    errors: pd.DataFrame
    suggestions: pd.DataFrame
    suggested_records: pd.DataFrame
    transferred_records: pd.DataFrame
    potential_conflicts: pd.DataFrame


def _is_missing(value: Any) -> bool:
    """Безопасно проверяет скалярное значение ячейки на пропуск."""

    if value is None:
        return True

    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        # Ячейка должна быть скаляром, но грязный источник не должен ломать парсер.
        return False


def clean_cell(value: Any) -> str:
    """Преобразует значение Excel-ячейки в чистую однострочную строку."""

    if _is_missing(value):
        return ""

    # Из-за NaN pandas часто превращает целые значения Excel в float (101 -> 101.0).
    if isinstance(value, Real) and not isinstance(value, bool):
        numeric_value = float(value)
        if math.isfinite(numeric_value) and numeric_value.is_integer():
            return str(int(numeric_value))

    text = str(value).replace("\r", " ").replace("\n", " ").replace("\xa0", " ")
    return WHITESPACE_PATTERN.sub(" ", text).strip()


def _clean_multiline_cell(value: Any) -> str:
    """Очищает текст, сохраняя смысловые переносы между занятиями."""

    if _is_missing(value):
        return ""
    lines = [
        WHITESPACE_PATTERN.sub(" ", line.replace("\xa0", " ")).strip()
        for line in str(value).replace("\r\n", "\n").replace("\r", "\n").split("\n")
    ]
    return "\n".join(line for line in lines if line)


def _excel_safe_text(value: Any, *, multiline: bool = False) -> str:
    """Не позволяет пользовательскому тексту стать формулой в XLSX."""

    text = _clean_multiline_cell(value) if multiline else clean_cell(value)
    return f"'{text}" if text.lstrip().startswith("=") else text


def _normalize_search_text(value: Any) -> str:
    """Нормализует регистр и русские е/ё только для сравнения строк."""

    return clean_cell(value).casefold().replace("ё", "е")


def _normalize_discipline(value: Any) -> str:
    """Нормализует название дисциплины для сверки разных Excel-источников."""

    text = WORKLOAD_SUBJECT_SUFFIX_PATTERN.sub("", clean_cell(value))
    normalized = _normalize_search_text(text)
    return " ".join(NAME_TOKEN_PATTERN.findall(normalized))


def _display_workload_discipline(value: Any) -> str:
    """Убирает из названия нагрузки часы, зачётные единицы и номер плана."""

    return clean_cell(WORKLOAD_SUBJECT_SUFFIX_PATTERN.sub("", clean_cell(value)))


def _normalize_lesson_type(value: Any) -> str:
    """Приводит обозначения лекций и практик к ЛК/ПР."""

    normalized = _normalize_search_text(value)
    if normalized.startswith("лк") or normalized.startswith("лек"):
        return "ЛК"
    if normalized.startswith("пр") or normalized.startswith("практ"):
        return "ПР"
    return clean_cell(value).upper()


def _teacher_surname_key(value: Any) -> str:
    """Возвращает нормализованную фамилию преподавателя."""

    tokens = NAME_TOKEN_PATTERN.findall(_normalize_search_text(value))
    return tokens[0] if tokens else ""


def _extract_groups(*values: Any) -> list[str]:
    """Извлекает названия групп из одного или нескольких грязных полей."""

    groups: list[str] = []
    for value in values:
        for group in GROUP_PATTERN.findall(clean_cell(value).upper()):
            if group not in groups:
                groups.append(group)
    return groups


def _parse_semester_number(value: Any) -> int | None:
    """Безопасно читает номер семестра из числа или строки."""

    if _is_missing(value):
        return None
    if isinstance(value, Real) and not isinstance(value, bool):
        numeric_value = float(value)
        if math.isfinite(numeric_value) and numeric_value.is_integer():
            return int(numeric_value)
    match = re.search(r"\d+", clean_cell(value))
    return int(match.group(0)) if match is not None else None


def parse_teacher_queries(value: str | Sequence[str]) -> list[str]:
    """Разбирает фамилии из строк, запятых и точек с запятой без дублей."""

    if isinstance(value, str):
        raw_queries = TEACHER_QUERY_SEPARATOR_PATTERN.split(value)
    else:
        raw_queries = list(value)

    queries: list[str] = []
    seen: set[str] = set()
    for raw_query in raw_queries:
        query = clean_cell(raw_query)
        normalized_query = _normalize_search_text(query)
        if not normalized_query or normalized_query in seen:
            continue
        seen.add(normalized_query)
        queries.append(query)

    return queries


def _format_pair(value: Any) -> str:
    """Убирает технический суффикс .0 у целочисленных номеров пар."""

    if _is_missing(value):
        return ""

    if isinstance(value, Real) and not isinstance(value, bool):
        numeric_value = float(value)
        if math.isfinite(numeric_value) and numeric_value.is_integer():
            return str(int(numeric_value))

    return clean_cell(value)


def _format_time(value: Any) -> str:
    """Нормализует время из строкового, datetime/time или Excel-числового вида."""

    if _is_missing(value):
        return ""

    if isinstance(value, (pd.Timestamp, datetime, time)):
        return value.strftime("%H:%M")

    # В старых .xls время иногда возвращается долей суток вместо datetime.time.
    if isinstance(value, Real) and not isinstance(value, bool):
        numeric_value = float(value)
        if math.isfinite(numeric_value) and 0 <= numeric_value < 1:
            total_minutes = round(numeric_value * 24 * 60) % (24 * 60)
            hours, minutes = divmod(total_minutes, 60)
            return f"{hours:02d}:{minutes:02d}"

    text_value = clean_cell(value)
    time_match = TIME_TEXT_PATTERN.fullmatch(text_value)
    if time_match is not None and time_match.group("seconds") in (None, "00"):
        return f"{int(time_match.group('hours')):02d}:{time_match.group('minutes')}"

    return text_value


def _build_time_range(start_value: Any, end_value: Any) -> str:
    """Собирает читаемый временной интервал без лишнего разделителя."""

    start = _format_time(start_value)
    end = _format_time(end_value)

    if start and end:
        return f"{start} – {end}"
    return start or end


def _is_time_value(value: Any) -> bool:
    """Проверяет, что значение похоже на время начала или окончания занятия."""

    if _is_missing(value):
        return False
    if isinstance(value, (pd.Timestamp, datetime, time)):
        return True
    if isinstance(value, Real) and not isinstance(value, bool):
        numeric_value = float(value)
        return math.isfinite(numeric_value) and 0 <= numeric_value < 1
    return TIME_TEXT_PATTERN.fullmatch(clean_cell(value)) is not None


def read_excel_file(file: bytes | bytearray | BinaryIO) -> pd.DataFrame:
    """Читает первый лист Excel без предположений о строке заголовков."""

    excel_file: BinaryIO
    if isinstance(file, (bytes, bytearray)):
        excel_file = BytesIO(file)
    else:
        excel_file = file

    try:
        excel_file.seek(0)
        dataframe = pd.read_excel(excel_file, header=None)
    except ImportError as exc:
        raise ScheduleReadError(
            "Не найден модуль для чтения Excel. Установите openpyxl для .xlsx "
            "и xlrd для .xls."
        ) from exc
    except Exception as exc:
        LOGGER.warning("Не удалось прочитать Excel-файл: %s", exc)
        raise ScheduleReadError(
            "Не удалось прочитать файл. Проверьте, что это неповреждённый "
            "Excel-файл формата .xls или .xlsx."
        ) from exc

    if dataframe.empty or dataframe.shape[1] == 0:
        raise ScheduleFormatError("Загруженный Excel-файл не содержит данных.")

    return dataframe


def read_excel_for_session(
    file_content: bytes,
    session_state: MutableMapping[str, Any],
) -> pd.DataFrame:
    """Читает файл один раз на сессию и переиспользует результат при rerun."""

    file_digest = sha256(file_content).hexdigest()
    cached_value = session_state.get(SESSION_CACHE_KEY)

    cache: dict[str, pd.DataFrame] = {}
    if isinstance(cached_value, dict):
        # Миграция кэша из старой версии, где хранился только один файл.
        old_digest = cached_value.get("digest")
        old_dataframe = cached_value.get("dataframe")
        if isinstance(old_digest, str) and isinstance(old_dataframe, pd.DataFrame):
            cache[old_digest] = old_dataframe
        else:
            cache = {
                digest: dataframe
                for digest, dataframe in cached_value.items()
                if isinstance(digest, str) and isinstance(dataframe, pd.DataFrame)
            }

    cached_dataframe = cache.get(file_digest)
    if cached_dataframe is not None:
        # Перемещаем недавно использованный файл в конец insertion-order кэша.
        cache.pop(file_digest)
        cache[file_digest] = cached_dataframe
        session_state[SESSION_CACHE_KEY] = cache
        return cached_dataframe

    dataframe = read_excel_file(file_content)
    cache[file_digest] = dataframe
    while len(cache) > MAX_CACHED_FILES:
        cache.pop(next(iter(cache)))
    session_state[SESSION_CACHE_KEY] = cache
    return dataframe


def read_schedule_records_for_session(
    file_content: bytes,
    session_state: MutableMapping[str, Any],
    file_digest: str | None = None,
) -> pd.DataFrame:
    """Кэширует уже распакованные строки расписания между rerun Streamlit."""

    digest = file_digest or sha256(file_content).hexdigest()
    cached_value = session_state.get(PARSED_SCHEDULE_CACHE_KEY)
    cache = (
        {
            cached_digest: dataframe
            for cached_digest, dataframe in cached_value.items()
            if isinstance(cached_digest, str) and isinstance(dataframe, pd.DataFrame)
        }
        if isinstance(cached_value, dict)
        else {}
    )

    cached_records = cache.get(digest)
    if cached_records is not None:
        cache.pop(digest)
        cache[digest] = cached_records
        session_state[PARSED_SCHEDULE_CACHE_KEY] = cache
        return cached_records

    source_df = read_excel_for_session(file_content, session_state)
    try:
        records = extract_schedule_records(source_df)
    finally:
        # После распаковки хранить одновременно исходную матрицу и плоские
        # записи нет смысла: на сервере это почти удваивает память каждой сессии.
        raw_cache = session_state.get(SESSION_CACHE_KEY)
        if isinstance(raw_cache, dict):
            raw_cache.pop(digest, None)
            session_state[SESSION_CACHE_KEY] = raw_cache
    cache[digest] = records
    while len(cache) > MAX_CACHED_FILES:
        cache.pop(next(iter(cache)))
    session_state[PARSED_SCHEDULE_CACHE_KEY] = cache
    return records


def find_group_row(df: pd.DataFrame) -> int:
    """Находит позицию строки с названиями групп среди первых десяти строк."""

    rows_to_scan = min(HEADER_SEARCH_ROWS, len(df.index))
    for row_idx in range(rows_to_scan):
        for value in df.iloc[row_idx].tolist():
            if GROUP_PATTERN.search(clean_cell(value)):
                return row_idx

    raise ScheduleFormatError(
        "Не найдена строка с названиями групп в первых 10 строках файла."
    )


def _prepare_schedule_rows(df: pd.DataFrame, group_row_idx: int) -> pd.DataFrame:
    """Выделяет строки расписания и распаковывает вертикальные объединения."""

    if df.shape[1] < COMMON_COLUMNS_COUNT:
        raise ScheduleFormatError(
            "В файле недостаточно столбцов для дня, пары, времени и недели."
        )

    data_start_idx = group_row_idx + 2
    if data_start_idx >= len(df.index):
        raise ScheduleFormatError(
            "После строки с группами не найдены строки расписания."
        )

    schedule_rows = df.iloc[data_start_idx:].copy()
    with pd.option_context("future.no_silent_downcasting", True):
        common_values = schedule_rows.iloc[:, :COMMON_COLUMNS_COUNT].ffill()
    schedule_rows.iloc[:, :COMMON_COLUMNS_COUNT] = common_values
    return schedule_rows


def _safe_cell(df: pd.DataFrame, row_idx: int, col_idx: int) -> str:
    """Читает ячейку по позиции, не падая на обрезанном хвосте таблицы."""

    if col_idx < 0 or col_idx >= df.shape[1]:
        return ""
    return clean_cell(df.iat[row_idx, col_idx])


def _pair_sort_key(value: Any) -> float:
    """Извлекает числовую часть пары; неизвестные значения отправляет в конец."""

    match = PAIR_NUMBER_PATTERN.search(clean_cell(value))
    if match is None:
        return math.inf

    try:
        return float(match.group(0).replace(",", "."))
    except ValueError:
        return math.inf


def _day_sort_key(value: Any) -> int:
    """Возвращает логический номер дня недели."""

    normalized_day = clean_cell(value).casefold().rstrip(".")
    return DAY_ORDER.get(normalized_day, len(DAY_ORDER))


def _is_schedule_row(df: pd.DataFrame, row_idx: int) -> bool:
    """Отделяет строки занятий от подписей и примечаний внизу Excel-листа."""

    day_order = _day_sort_key(df.iat[row_idx, 0])
    pair_order = _pair_sort_key(df.iat[row_idx, 1])
    week = clean_cell(df.iat[row_idx, 4]).casefold()

    return (
        day_order <= 5
        and math.isfinite(pair_order)
        and _is_time_value(df.iat[row_idx, 2])
        and _is_time_value(df.iat[row_idx, 3])
        and week in {"i", "ii"}
    )


def sort_schedule(result: pd.DataFrame) -> pd.DataFrame:
    """Сортирует результат по дню и номеру пары, сохраняя порядок совпадений."""

    if result.empty:
        return result.reindex(columns=OUTPUT_COLUMNS).reset_index(drop=True)

    sortable = result.copy()
    sortable["__day_order"] = sortable["День недели"].map(_day_sort_key)
    sortable["__pair_order"] = sortable["Пара"].map(_pair_sort_key)
    sortable["__source_order"] = range(len(sortable.index))

    sortable = sortable.sort_values(
        by=["__day_order", "__pair_order", "__source_order"],
        kind="mergesort",
        na_position="last",
    )
    return sortable.drop(
        columns=["__day_order", "__pair_order", "__source_order"]
    ).reset_index(drop=True)


def sort_internal_schedule(
    result: pd.DataFrame,
    teacher_queries: Sequence[str],
) -> pd.DataFrame:
    """Сортирует пакетный результат с сохранением порядка введённых фамилий."""

    if result.empty:
        return result.reindex(columns=INTERNAL_COLUMNS).reset_index(drop=True)

    query_order = {
        _normalize_search_text(query): position
        for position, query in enumerate(teacher_queries)
    }
    sortable = result.copy()
    sortable["__query_order"] = sortable[QUERY_COLUMN].map(
        lambda value: query_order.get(_normalize_search_text(value), len(query_order))
    )
    sortable["__day_order"] = sortable["День недели"].map(_day_sort_key)
    sortable["__pair_order"] = sortable["Пара"].map(_pair_sort_key)
    sortable["__week_order"] = sortable["Неделя"].map(
        lambda value: 0 if clean_cell(value).casefold() == "i" else 1
    )
    sortable["__source_order"] = range(len(sortable.index))
    sortable = sortable.sort_values(
        by=[
            "__query_order",
            "__day_order",
            "__pair_order",
            "__week_order",
            "__source_order",
        ],
        kind="mergesort",
        na_position="last",
    )
    return sortable.drop(
        columns=[
            "__query_order",
            "__day_order",
            "__pair_order",
            "__week_order",
            "__source_order",
        ]
    ).reset_index(drop=True)


def _join_unique(values: Sequence[Any], separator: str = ", ") -> str:
    """Объединяет непустые значения без дублей, сохраняя исходный порядок."""

    unique_values: list[str] = []
    for value in values:
        cleaned_value = clean_cell(value)
        if cleaned_value and cleaned_value not in unique_values:
            unique_values.append(cleaned_value)
    return separator.join(unique_values)


def _lesson_count_text(count: int) -> str:
    """Возвращает согласованную русскую подпись количества занятий."""

    last_two_digits = count % 100
    last_digit = count % 10
    if 11 <= last_two_digits <= 14:
        noun = "занятий"
    elif last_digit == 1:
        noun = "занятие"
    elif 2 <= last_digit <= 4:
        noun = "занятия"
    else:
        noun = "занятий"
    return f"{count} {noun}"


def _teacher_conflict_identity(query: Any, teacher: Any) -> str:
    """Определяет преподавателя для сравнения занятий одного временного слота."""

    normalized_query = _normalize_search_text(query)
    if len(normalized_query.split()) >= 2:
        # Полное имя или фамилия с инициалами надёжнее плавающей записи в Excel.
        return normalized_query
    return _normalize_search_text(teacher) or normalized_query


def collapse_schedule_conflicts(records: pd.DataFrame) -> pd.DataFrame:
    """Сворачивает занятия слота и отмечает только разные занятия преподавателя."""

    if records.empty:
        return pd.DataFrame(columns=DISPLAY_COLUMNS)

    required_columns = set(INTERNAL_COLUMNS)
    if missing_columns := required_columns.difference(records.columns):
        raise ValueError(
            "Для поиска накладок не хватает столбцов: "
            + ", ".join(sorted(missing_columns))
        )

    display_records: list[dict[str, str]] = []
    grouped_records = records.copy()
    identity_column = "__teacher_conflict_identity"
    grouped_records[identity_column] = [
        _teacher_conflict_identity(query, teacher)
        for query, teacher in zip(
            grouped_records[QUERY_COLUMN],
            grouped_records[TEACHER_COLUMN],
            strict=True,
        )
    ]
    slot_columns = [
        QUERY_COLUMN,
        identity_column,
        "День недели",
        "Пара",
        "Неделя",
    ]

    for _, slot_records in grouped_records.groupby(
        slot_columns,
        sort=False,
        dropna=False,
    ):
        logical_lessons: dict[
            tuple[str, str, str, str, str],
            dict[str, Any],
        ] = {}
        for record in slot_records.to_dict(orient="records"):
            lesson_values = (
                clean_cell(record["Время"]),
                clean_cell(record["Дисциплина"]),
                clean_cell(record["Вид занятий"]),
                clean_cell(record["Аудитория"]),
            )
            lesson_type_key = _normalize_lesson_type(lesson_values[2])
            discipline_key = _normalize_room_conflict_discipline(lesson_values[1])
            group = clean_cell(record["Группа"])
            # Несколько групп одной лекции — одно занятие. Практики разных
            # групп являются отдельными занятиями даже при одинаковом предмете.
            non_lecture_group_key = "" if lesson_type_key == "ЛК" else group.upper()
            lesson_key = (
                _normalize_search_text(lesson_values[0]),
                discipline_key,
                _normalize_search_text(lesson_values[2]),
                _normalize_search_text(lesson_values[3]),
                non_lecture_group_key,
            )
            lesson = logical_lessons.setdefault(
                lesson_key,
                {
                    "time": lesson_values[0],
                    "discipline": lesson_values[1],
                    "lesson_type": lesson_values[2],
                    "auditorium": lesson_values[3],
                    "suggestion_labels": [],
                    "groups": [],
                    "sources": [],
                    "weeks": set(),
                },
            )
            lesson["weeks"].update(
                _extract_lesson_weeks(lesson_values[1], record["Неделя"])
            )
            suggestion_label = clean_cell(record[SUGGESTED_COLUMN])
            if suggestion_label and suggestion_label not in lesson["suggestion_labels"]:
                lesson["suggestion_labels"].append(suggestion_label)
            source = clean_cell(record[SOURCE_COLUMN])
            if group and group not in lesson["groups"]:
                lesson["groups"].append(group)
            if source and source not in lesson["sources"]:
                lesson["sources"].append(source)

        lessons = list(logical_lessons.values())
        first_record = slot_records.iloc[0]
        all_weeks = sorted({week for lesson in lessons for week in lesson["weeks"]})
        simultaneous_counts = {
            week: sum(week in lesson["weeks"] for lesson in lessons)
            for week in all_weeks
        }
        maximum_simultaneous = max(simultaneous_counts.values(), default=0)
        conflict_weeks = [
            week for week, count in simultaneous_counts.items() if count > 1
        ]
        if maximum_simultaneous > 1:
            normalized_week = _normalize_search_text(first_record["Неделя"])
            default_weeks = (
                set(range(1, 19, 2)) if normalized_week == "i" else set(range(2, 19, 2))
            )
            weeks_suffix = (
                ""
                if set(conflict_weeks) == default_weeks
                else f" (учебные недели: {_format_week_numbers(conflict_weeks)})"
            )
            conflict_text = (
                f"⚠ {_lesson_count_text(maximum_simultaneous)} одновременно"
                f"{weeks_suffix}"
            )
        else:
            conflict_text = ""
        display_records.append(
            {
                QUERY_COLUMN: clean_cell(first_record[QUERY_COLUMN]),
                TEACHER_COLUMN: _join_unique(
                    slot_records[TEACHER_COLUMN].tolist(),
                    separator="\n",
                ),
                SUGGESTED_COLUMN: _join_unique(
                    [
                        label
                        for lesson in lessons
                        for label in lesson["suggestion_labels"]
                    ],
                    separator="\n",
                ),
                "День недели": clean_cell(first_record["День недели"]),
                "Пара": clean_cell(first_record["Пара"]),
                "Время": _join_unique(
                    [lesson["time"] for lesson in lessons],
                    separator="\n",
                ),
                "Неделя": clean_cell(first_record["Неделя"]),
                CONFLICT_COLUMN: conflict_text,
                "Группа": "\n".join(
                    _join_unique(lesson["groups"]) for lesson in lessons
                ),
                "Дисциплина": "\n".join(lesson["discipline"] for lesson in lessons),
                "Вид занятий": "\n".join(lesson["lesson_type"] for lesson in lessons),
                "Аудитория": "\n".join(lesson["auditorium"] for lesson in lessons),
                SOURCE_COLUMN: "\n".join(
                    _join_unique(lesson["sources"]) for lesson in lessons
                ),
            }
        )

    display_frame = pd.DataFrame.from_records(
        display_records,
        columns=DISPLAY_COLUMNS,
    )
    return sort_internal_schedule(
        display_frame,
        parse_teacher_queries(records[QUERY_COLUMN].tolist()),
    )


def _is_physical_room(value: Any) -> bool:
    """Отбрасывает пустые и дистанционные значения аудитории."""

    room = _normalize_search_text(value)
    if not room or room in {"-", "—"}:
        return False
    if room in NON_SPECIFIC_ROOM_VALUES:
        return False
    return not any(marker in room for marker in EXCLUDED_ROOM_MARKERS)


def _normalize_room_conflict_discipline(value: Any) -> str:
    """Убирает номера недель перед сравнением названий общих лекций."""

    without_weeks = WEEK_SEQUENCE_PATTERN.sub(" ", clean_cell(value))
    return _normalize_discipline(without_weeks)


def _extract_lesson_weeks(discipline: Any, week_type: Any) -> frozenset[int]:
    """Извлекает фактические недели занятия или возвращает недели I/II."""

    weeks: set[int] = set()
    for match in WEEK_SEQUENCE_PATTERN.finditer(clean_cell(discipline)):
        for part in re.split(r"\s*[,;]\s*", match.group("weeks")):
            range_match = re.fullmatch(r"(\d{1,2})\s*[-–—]\s*(\d{1,2})", part)
            if range_match is None:
                week_number = int(part)
                if 1 <= week_number <= 20:
                    weeks.add(week_number)
                continue
            range_start, range_end = map(int, range_match.groups())
            if range_start > range_end:
                range_start, range_end = range_end, range_start
            weeks.update(range(max(1, range_start), min(20, range_end) + 1))

    if weeks:
        return frozenset(weeks)

    normalized_week = _normalize_search_text(week_type)
    if normalized_week == "i":
        return frozenset(range(1, 19, 2))
    if normalized_week == "ii":
        return frozenset(range(2, 19, 2))
    return frozenset()


def _format_week_numbers(weeks: Sequence[int]) -> str:
    """Форматирует номера фактических учебных недель по возрастанию."""

    return ", ".join(str(week) for week in sorted(set(weeks)))


def find_room_conflicts(
    schedule_records: pd.DataFrame,
    teacher_queries: Sequence[str],
) -> pd.DataFrame:
    """Ищет двойное бронирование кабинетов для выбранных преподавателей.

    Сравнение выполняется по всему загруженному расписанию. Несколько групп
    одной и той же лекции в одном кабинете считаются одним занятием.
    """

    queries = parse_teacher_queries(teacher_queries)
    if not queries or schedule_records.empty:
        return pd.DataFrame(columns=ROOM_CONFLICT_COLUMNS)
    normalized_queries = [(query, _normalize_search_text(query)) for query in queries]

    schedule = schedule_records.reindex(columns=ALL_SCHEDULE_COLUMNS).copy()
    schedule = schedule[schedule["Аудитория"].map(_is_physical_room)].copy()
    if schedule.empty:
        return pd.DataFrame(columns=ROOM_CONFLICT_COLUMNS)

    schedule["__day_key"] = schedule["День недели"].map(_day_sort_key)
    schedule["__pair_key"] = schedule["Пара"].map(_format_pair)
    schedule["__room_key"] = schedule["Аудитория"].map(_normalize_search_text)
    schedule = schedule[
        (schedule["__day_key"] <= 5)
        & (schedule["__pair_key"] != "")
        & (schedule["Неделя"].map(_normalize_search_text) != "")
    ]
    schedule["__teacher_key"] = schedule[TEACHER_COLUMN].map(_normalize_search_text)
    selected_teacher_mask = schedule["__teacher_key"].map(
        lambda teacher: any(
            normalized_query in teacher for _, normalized_query in normalized_queries
        )
    )
    relevant_slots = set(
        schedule.loc[
            selected_teacher_mask,
            ["__day_key", "__pair_key", "__room_key"],
        ].itertuples(index=False, name=None)
    )
    if not relevant_slots:
        return pd.DataFrame(columns=ROOM_CONFLICT_COLUMNS)
    schedule = schedule[
        [
            slot in relevant_slots
            for slot in schedule[["__day_key", "__pair_key", "__room_key"]].itertuples(
                index=False, name=None
            )
        ]
    ]

    conflict_records: list[dict[str, str]] = []
    slot_columns = ["__day_key", "__pair_key", "__room_key"]
    for _, slot_records in schedule.groupby(
        slot_columns,
        sort=False,
        dropna=False,
    ):
        occupancies: dict[tuple[str, ...], dict[str, Any]] = {}
        for record_index, record in enumerate(slot_records.to_dict(orient="records")):
            discipline = clean_cell(record["Дисциплина"])
            discipline_key = _normalize_room_conflict_discipline(discipline)
            lesson_type = _normalize_lesson_type(record["Вид занятий"])
            group = clean_cell(record["Группа"]).upper()
            teacher = clean_cell(record[TEACHER_COLUMN])
            source = clean_cell(record[SOURCE_COLUMN])

            if lesson_type == "ЛК" and discipline_key:
                # Общая лекция одного предмета может занимать одну аудиторию
                # сразу для нескольких групп и преподавателей.
                occupancy_key = ("lecture", discipline_key)
            else:
                group_identity = group or _teacher_surname_key(teacher)
                occupancy_key = (
                    "lesson",
                    discipline_key or f"row-{record_index}",
                    lesson_type,
                    group_identity or source or f"row-{record_index}",
                )

            occupancy = occupancies.setdefault(
                occupancy_key,
                {
                    "teachers": [],
                    "groups": [],
                    "disciplines": [],
                    "lesson_types": [],
                    "sources": [],
                    "times": [],
                    "week_types": [],
                    "weeks": set(),
                },
            )
            for field, value in (
                ("teachers", teacher),
                ("groups", group),
                ("disciplines", discipline),
                ("lesson_types", clean_cell(record["Вид занятий"])),
                ("sources", source),
                ("times", clean_cell(record["Время"])),
                ("week_types", clean_cell(record["Неделя"])),
            ):
                if value and value not in occupancy[field]:
                    occupancy[field].append(value)
            occupancy["weeks"].update(
                _extract_lesson_weeks(discipline, record["Неделя"])
            )

        if len(occupancies) <= 1:
            continue

        conflicting_signatures: dict[tuple[tuple[str, ...], ...], list[int]] = {}
        occupancy_keys = list(occupancies)
        all_weeks = sorted(
            {week for occupancy in occupancies.values() for week in occupancy["weeks"]}
        )
        for week in all_weeks:
            signature = tuple(
                occupancy_key
                for occupancy_key in occupancy_keys
                if week in occupancies[occupancy_key]["weeks"]
            )
            if len(signature) > 1:
                conflicting_signatures.setdefault(signature, []).append(week)

        first = slot_records.iloc[0]
        for signature, conflict_weeks in conflicting_signatures.items():
            logical_lessons = [occupancies[key] for key in signature]
            matching_teacher_values = [
                teacher for lesson in logical_lessons for teacher in lesson["teachers"]
            ]
            matched_queries = [
                query
                for query, normalized_query in normalized_queries
                if any(
                    normalized_query in _normalize_search_text(teacher)
                    for teacher in matching_teacher_values
                )
            ]
            if not matched_queries:
                continue

            conflict_records.append(
                {
                    QUERY_COLUMN: ", ".join(matched_queries),
                    "День недели": clean_cell(first["День недели"]),
                    "Пара": clean_cell(first["Пара"]),
                    "Время": _join_unique(
                        [
                            lesson_time
                            for lesson in logical_lessons
                            for lesson_time in lesson["times"]
                        ],
                        "\n",
                    ),
                    "Неделя": _join_unique(
                        [
                            week_type
                            for lesson in logical_lessons
                            for week_type in lesson["week_types"]
                        ]
                    ),
                    "Учебные недели": _format_week_numbers(conflict_weeks),
                    "Аудитория": clean_cell(first["Аудитория"]),
                    ROOM_CONFLICT_COLUMN: (
                        f"⚠ {_lesson_count_text(len(logical_lessons))} "
                        "в одной аудитории"
                    ),
                    "Преподаватели": "\n".join(
                        _join_unique(lesson["teachers"]) for lesson in logical_lessons
                    ),
                    "Группы": "\n".join(
                        _join_unique(lesson["groups"]) for lesson in logical_lessons
                    ),
                    "Дисциплины": "\n".join(
                        _join_unique(lesson["disciplines"])
                        for lesson in logical_lessons
                    ),
                    "Виды занятий": "\n".join(
                        _join_unique(lesson["lesson_types"])
                        for lesson in logical_lessons
                    ),
                    SOURCE_COLUMN: "\n".join(
                        _join_unique(lesson["sources"]) for lesson in logical_lessons
                    ),
                }
            )

    conflicts = pd.DataFrame.from_records(
        conflict_records,
        columns=ROOM_CONFLICT_COLUMNS,
    )
    if conflicts.empty:
        return conflicts

    query_order = {
        _normalize_search_text(query): position
        for position, query in enumerate(queries)
    }
    sortable = conflicts.copy()
    sortable["__query_order"] = sortable[QUERY_COLUMN].map(
        lambda value: min(
            (
                query_order[_normalize_search_text(query)]
                for query in parse_teacher_queries(value)
                if _normalize_search_text(query) in query_order
            ),
            default=len(query_order),
        )
    )
    sortable["__day_order"] = sortable["День недели"].map(_day_sort_key)
    sortable["__pair_order"] = sortable["Пара"].map(_pair_sort_key)
    sortable["__week_order"] = sortable["Учебные недели"].map(
        lambda value: min(
            (int(part) for part in clean_cell(value).split(", ") if part.isdigit()),
            default=99,
        )
    )
    return (
        sortable.sort_values(
            by=[
                "__query_order",
                "__day_order",
                "__pair_order",
                "__week_order",
                "Аудитория",
            ],
            kind="mergesort",
        )
        .drop(
            columns=[
                "__query_order",
                "__day_order",
                "__pair_order",
                "__week_order",
            ]
        )
        .reset_index(drop=True)
    )


def extract_schedule_records(
    df: pd.DataFrame,
    source_name: str = "",
) -> pd.DataFrame:
    """Извлекает все заполненные занятия, включая строки без преподавателя."""

    group_row_idx = find_group_row(df)
    schedule_rows = _prepare_schedule_rows(df, group_row_idx)

    if df.shape[1] <= FIRST_GROUP_COLUMN + 2:
        raise ScheduleFormatError("В файле не найдены столбцы групп и преподавателей.")

    records: list[dict[str, str]] = []
    group_columns = range(
        FIRST_GROUP_COLUMN,
        df.shape[1] - 2,
        GROUP_BLOCK_SIZE,
    )

    for row_idx in range(len(schedule_rows.index)):
        if not _is_schedule_row(schedule_rows, row_idx):
            continue

        day = _safe_cell(schedule_rows, row_idx, 0)
        pair = _format_pair(schedule_rows.iat[row_idx, 1])
        lesson_time = _build_time_range(
            schedule_rows.iat[row_idx, 2],
            schedule_rows.iat[row_idx, 3],
        )
        week = _safe_cell(schedule_rows, row_idx, 4)

        for col_idx in group_columns:
            group = _safe_cell(df, group_row_idx, col_idx)
            if not group or GROUP_PATTERN.search(group) is None:
                continue

            discipline = _safe_cell(schedule_rows, row_idx, col_idx)
            lesson_type = _safe_cell(schedule_rows, row_idx, col_idx + 1)
            teacher = _safe_cell(schedule_rows, row_idx, col_idx + 2)
            auditorium = _safe_cell(schedule_rows, row_idx, col_idx + 3)
            if not any((discipline, lesson_type, teacher, auditorium)):
                continue
            records.append(
                {
                    TEACHER_COLUMN: teacher,
                    "День недели": day,
                    "Пара": pair,
                    "Время": lesson_time,
                    "Неделя": week,
                    "Группа": group,
                    "Дисциплина": discipline,
                    "Вид занятий": lesson_type,
                    "Аудитория": auditorium,
                    SOURCE_COLUMN: clean_cell(source_name),
                }
            )

    result = pd.DataFrame.from_records(records, columns=ALL_SCHEDULE_COLUMNS)
    if result.empty:
        return result

    sortable = result.copy()
    sortable["__day_order"] = sortable["День недели"].map(_day_sort_key)
    sortable["__pair_order"] = sortable["Пара"].map(_pair_sort_key)
    sortable["__week_order"] = sortable["Неделя"].map(
        lambda value: 0 if _normalize_search_text(value) == "i" else 1
    )
    sortable["__source_order"] = range(len(sortable.index))
    sortable = sortable.sort_values(
        by=["__day_order", "__pair_order", "__week_order", "__source_order"],
        kind="mergesort",
        na_position="last",
    )
    return sortable.drop(
        columns=["__day_order", "__pair_order", "__week_order", "__source_order"]
    ).reset_index(drop=True)


def _filter_schedule_records(
    records: pd.DataFrame,
    teacher_queries: Sequence[str],
) -> pd.DataFrame:
    """Фильтрует полный набор занятий по введённым фамилиям или частям ФИО."""

    queries = parse_teacher_queries(teacher_queries)
    if not queries or records.empty:
        return pd.DataFrame(columns=INTERNAL_COLUMNS)

    normalized_queries = [(query, _normalize_search_text(query)) for query in queries]
    filtered_records: list[dict[str, str]] = []
    for record in records.to_dict(orient="records"):
        normalized_teacher = _normalize_search_text(record[TEACHER_COLUMN])
        for query, normalized_query in normalized_queries:
            if normalized_query not in normalized_teacher:
                continue
            filtered_records.append({QUERY_COLUMN: query, **record})

    for record in filtered_records:
        record[SUGGESTED_COLUMN] = ""

    result = pd.DataFrame.from_records(filtered_records, columns=INTERNAL_COLUMNS)
    return sort_internal_schedule(result, queries)


def parse_schedule_multi(
    df: pd.DataFrame,
    teacher_queries: Sequence[str],
    source_name: str = "",
) -> pd.DataFrame:
    """Извлекает занятия нескольких преподавателей за один проход по листу."""

    queries = parse_teacher_queries(teacher_queries)
    if not queries:
        raise ValueError("Введите хотя бы одну фамилию преподавателя.")
    return _filter_schedule_records(
        extract_schedule_records(df, source_name),
        queries,
    )


def parse_schedule(df: pd.DataFrame, teacher_query: str) -> pd.DataFrame:
    """Совместимый однопользовательский интерфейс старого парсера."""

    result = parse_schedule_multi(df, [teacher_query])
    return sort_schedule(result.reindex(columns=OUTPUT_COLUMNS))


def _deduplicate_batch_records(
    records: pd.DataFrame,
    teacher_queries: Sequence[str],
) -> pd.DataFrame:
    """Удаляет полные дубли занятий, сохраняя разные группы и запросы."""

    if records.empty:
        return records.reindex(columns=INTERNAL_COLUMNS).reset_index(drop=True)

    deduplication_columns = [
        QUERY_COLUMN,
        TEACHER_COLUMN,
        *OUTPUT_COLUMNS,
    ]
    deduplicated = records.drop_duplicates(
        subset=deduplication_columns,
        keep="first",
        ignore_index=True,
    )
    return sort_internal_schedule(deduplicated, teacher_queries)


def _deduplicate_all_schedule_records(records: pd.DataFrame) -> pd.DataFrame:
    """Удаляет полные дубли из набора всех занятий для сверки нагрузки."""

    if records.empty:
        return records.reindex(columns=ALL_SCHEDULE_COLUMNS).reset_index(drop=True)
    return records.drop_duplicates(
        subset=[TEACHER_COLUMN, *OUTPUT_COLUMNS],
        keep="first",
        ignore_index=True,
    ).reindex(columns=ALL_SCHEDULE_COLUMNS)


def parse_schedule_files(
    files: Sequence[tuple[str, bytes]],
    teacher_queries: Sequence[str],
    session_state: MutableMapping[str, Any],
) -> BatchParseResult:
    """Продолжает пакетную обработку, даже если отдельный файл повреждён."""

    queries = parse_teacher_queries(teacher_queries)
    processed_files: list[str] = []
    duplicate_files: list[str] = []
    errors: list[BatchFileError] = []
    dataframes: list[pd.DataFrame] = []
    all_dataframes: list[pd.DataFrame] = []
    seen_digests: set[str] = set()

    for position, (raw_filename, file_content) in enumerate(files, start=1):
        filename = clean_cell(raw_filename) or f"Файл {position}"
        file_digest = sha256(file_content).hexdigest()
        if file_digest in seen_digests:
            duplicate_files.append(filename)
            continue
        seen_digests.add(file_digest)

        try:
            cached_records = read_schedule_records_for_session(
                file_content,
                session_state,
                file_digest,
            )
            all_parsed = cached_records.assign(**{SOURCE_COLUMN: filename})
            parsed = _filter_schedule_records(all_parsed, queries)
        except ScheduleError as exc:
            errors.append(BatchFileError(filename, str(exc)))
            continue
        except Exception:
            LOGGER.exception("Непредвиденная ошибка файла %s", filename)
            errors.append(
                BatchFileError(
                    filename,
                    "Непредвиденная ошибка обработки файла.",
                )
            )
            continue

        processed_files.append(filename)
        dataframes.append(parsed)
        all_dataframes.append(all_parsed)

    if dataframes:
        combined = pd.concat(dataframes, ignore_index=True)
        records = _deduplicate_batch_records(combined, queries)
    else:
        records = pd.DataFrame(columns=INTERNAL_COLUMNS)

    if all_dataframes:
        all_records = _deduplicate_all_schedule_records(
            pd.concat(all_dataframes, ignore_index=True)
        )
    else:
        all_records = pd.DataFrame(columns=ALL_SCHEDULE_COLUMNS)

    return BatchParseResult(
        records=records,
        all_records=all_records,
        processed_files=tuple(processed_files),
        duplicate_files=tuple(duplicate_files),
        errors=tuple(errors),
    )


def _find_workload_columns(df: pd.DataFrame) -> tuple[int, dict[str, int]]:
    """Находит строку заголовков и обязательные столбцы файла нагрузки."""

    required_headers = {
        "subject": "мероприятие реестра",
        "lesson_type": "вид потока",
        "plan_group": "план. поток",
        "group": "группа",
        "semester": "семестр",
        "teacher": "ппс",
    }
    for row_idx in range(min(15, len(df.index))):
        normalized_headers = [
            _normalize_search_text(value) for value in df.iloc[row_idx].tolist()
        ]
        columns: dict[str, int] = {}
        for key, expected_header in required_headers.items():
            for column_idx, header in enumerate(normalized_headers):
                if header == expected_header or (
                    key == "subject" and header.startswith(expected_header)
                ):
                    columns[key] = column_idx
                    break
        if len(columns) == len(required_headers):
            return row_idx, columns

    raise WorkloadFormatError(
        "В файле нагрузки не найдены столбцы «Мероприятие реестра», "
        "«Вид потока», «Группа», «Семестр» и «ППС»."
    )


def parse_workload_file(
    file: bytes | bytearray | BinaryIO,
    academic_term: str,
) -> pd.DataFrame:
    """Читает нагрузку и оставляет ЛК/ПР нужной чётности семестра."""

    df = read_excel_file(file)
    header_row_idx, columns = _find_workload_columns(df)
    normalized_term = _normalize_search_text(academic_term)
    if normalized_term not in {"осень", "весна"}:
        raise ValueError("Учебный семестр должен быть «Осень» или «Весна».")
    expected_parity = 1 if normalized_term == "осень" else 0

    records: list[dict[str, Any]] = []
    for row_idx in range(header_row_idx + 1, len(df.index)):
        raw_subject = df.iat[row_idx, columns["subject"]]
        discipline = _display_workload_discipline(raw_subject)
        lesson_type = _normalize_lesson_type(df.iat[row_idx, columns["lesson_type"]])
        semester = _parse_semester_number(df.iat[row_idx, columns["semester"]])
        if (
            not discipline
            or lesson_type not in WORKLOAD_CLASS_TYPES
            or semester is None
            or semester % 2 != expected_parity
        ):
            continue

        teacher = clean_cell(df.iat[row_idx, columns["teacher"]])
        groups = _extract_groups(
            df.iat[row_idx, columns["plan_group"]],
            df.iat[row_idx, columns["group"]],
        ) or [""]
        for group in groups:
            records.append(
                {
                    "Преподаватель нагрузки": teacher,
                    "Дисциплина": discipline,
                    "Вид занятий": lesson_type,
                    "Группа": group,
                    "Семестр": semester,
                    "Строка нагрузки": row_idx + 1,
                }
            )

    workload = pd.DataFrame.from_records(records, columns=WORKLOAD_COLUMNS)
    if workload.empty:
        term_label = "нечётных" if expected_parity else "чётных"
        raise WorkloadFormatError(
            f"В нагрузке не найдены ЛК/ПР для {term_label} семестров."
        )
    return workload.drop_duplicates(
        subset=[
            "Преподаватель нагрузки",
            "Дисциплина",
            "Вид занятий",
            "Группа",
            "Семестр",
        ],
        keep="first",
        ignore_index=True,
    )


def _schedule_place(records: pd.DataFrame) -> str:
    """Формирует краткое описание возможного места в расписании."""

    places: list[str] = []
    for record in records.to_dict(orient="records"):
        details = [clean_cell(record["Группа"])]
        auditorium = clean_cell(record["Аудитория"])
        if auditorium:
            details.append(f"ауд. {auditorium}")
        details.append(clean_cell(record[SOURCE_COLUMN]))
        place = (
            f"{clean_cell(record['День недели'])}, "
            f"пара {clean_cell(record['Пара'])}, "
            f"неделя {clean_cell(record['Неделя'])}, "
            f"{clean_cell(record['Время'])}; "
            f"{'; '.join(detail for detail in details if detail)}"
        )
        if place not in places:
            places.append(place)
    return "\n".join(places)


def _collapse_schedule_place_values(values: Sequence[Any]) -> str:
    """Объединяет группы в описаниях одного временного слота."""

    slot_groups: dict[tuple[str, tuple[str, ...]], list[str]] = {}
    unparsed_lines: list[str] = []
    for value in values:
        for line in _clean_multiline_cell(value).splitlines():
            parts = [part.strip() for part in line.split(";")]
            if len(parts) < 3:
                if line and line not in unparsed_lines:
                    unparsed_lines.append(line)
                continue
            key = (parts[0], tuple(parts[2:]))
            groups = slot_groups.setdefault(key, [])
            group = parts[1]
            if group and group not in groups:
                groups.append(group)

    collapsed = [
        "; ".join((slot, ", ".join(groups), *suffix))
        for (slot, suffix), groups in slot_groups.items()
    ]
    return "\n".join([*collapsed, *unparsed_lines])


def _collapse_workload_issue_records(
    records: Sequence[dict[str, str]],
) -> pd.DataFrame:
    """Сворачивает замечания до одной строки исходного потока нагрузки."""

    if not records:
        return pd.DataFrame(columns=WORKLOAD_ISSUE_COLUMNS)

    frame = pd.DataFrame.from_records(records)
    group_columns = [
        WORKLOAD_ROWS_KEY_COLUMN,
        "Преподаватели по нагрузке",
        "Дисциплина",
        "Вид занятий",
        "Семестр",
    ]
    collapsed_records: list[dict[str, str]] = []
    for _, issues in frame.groupby(group_columns, sort=False, dropna=False):
        first = issues.iloc[0]
        collapsed_records.append(
            {
                "Статус": _join_unique(issues["Статус"].tolist(), separator="\n"),
                "Проблема": _join_unique(
                    issues["Проблема"].tolist(),
                    separator="\n",
                ),
                "Преподаватели по нагрузке": clean_cell(
                    first["Преподаватели по нагрузке"]
                ),
                "Дисциплина": clean_cell(first["Дисциплина"]),
                "Вид занятий": clean_cell(first["Вид занятий"]),
                "Группа": _join_unique(
                    issues["Группа"].tolist(),
                    separator=", ",
                ),
                "Семестр": clean_cell(first["Семестр"]),
                "ФИО в расписании": _join_unique(
                    issues["ФИО в расписании"].tolist(),
                    separator="\n",
                ),
                "Предлагаемое ФИО": _join_unique(
                    issues["Предлагаемое ФИО"].tolist(),
                    separator="\n",
                ),
                "Возможная накладка": _join_unique(
                    issues["Возможная накладка"].tolist(),
                    separator="\n",
                ),
                "Возможное место в расписании": _collapse_schedule_place_values(
                    issues["Возможное место в расписании"].tolist()
                ),
            }
        )
    return pd.DataFrame.from_records(
        collapsed_records,
        columns=WORKLOAD_ISSUE_COLUMNS,
    )


def _suggested_teacher_conflicts(
    blank_candidates: pd.DataFrame,
    schedule: pd.DataFrame,
    suggested_teacher: str,
    suggested_discipline_key: str,
) -> str:
    """Ищет занятия, конфликтующие с предлагаемой подстановкой ФИО."""

    surname = _teacher_surname_key(suggested_teacher)
    if not surname:
        return ""
    teacher_schedule = schedule[schedule["__teacher_surname"] == surname]
    conflicts: list[str] = []
    for candidate in blank_candidates.to_dict(orient="records"):
        candidate_weeks = _extract_lesson_weeks(
            candidate["Дисциплина"],
            candidate["Неделя"],
        )
        same_slot = teacher_schedule[
            (teacher_schedule["__day_key"] == _day_sort_key(candidate["День недели"]))
            & (teacher_schedule["__pair_key"] == _format_pair(candidate["Пара"]))
            & (teacher_schedule["__discipline_key"] != suggested_discipline_key)
        ]
        for conflict in same_slot.to_dict(orient="records"):
            common_weeks = candidate_weeks.intersection(conflict["__lesson_weeks"])
            if not common_weeks:
                continue
            description = (
                f"{clean_cell(conflict['День недели'])}, "
                f"пара {clean_cell(conflict['Пара'])}, "
                f"учебные недели {_format_week_numbers(common_weeks)}: "
                f"{clean_cell(conflict['Дисциплина'])} "
                f"({clean_cell(conflict['Группа'])})"
            )
            if description not in conflicts:
                conflicts.append(description)
    return "\n".join(conflicts)


def _find_transferred_workload_issues(
    workload: pd.DataFrame,
    schedule: pd.DataFrame,
    normalized_queries: Sequence[tuple[str, str]],
) -> tuple[list[dict[str, str]], list[dict[str, str]], set[str]]:
    """Находит занятия, оставшиеся в расписании после передачи нагрузки."""

    assignment_columns = [
        "__discipline_key",
        "__lesson_type_key",
        "__group_key",
    ]
    workload_by_assignment = {
        tuple(clean_cell(value) for value in key): assignment
        for key, assignment in workload.groupby(
            assignment_columns,
            sort=False,
            dropna=False,
        )
    }
    selected_schedule = schedule[
        schedule["__lesson_type_key"].isin(WORKLOAD_CLASS_TYPES)
        & schedule[TEACHER_COLUMN].map(
            lambda teacher: any(
                normalized_query in _normalize_search_text(teacher)
                for _, normalized_query in normalized_queries
            )
        )
    ]

    issues: list[dict[str, str]] = []
    transferred_records: list[dict[str, str]] = []
    transferred_workload_rows: set[str] = set()
    schedule_columns = [*assignment_columns, TEACHER_COLUMN]
    for key, candidates in selected_schedule.groupby(
        schedule_columns,
        sort=False,
        dropna=False,
    ):
        discipline_key, lesson_type, group, scheduled_teacher = (
            clean_cell(value) for value in key
        )
        assignment = workload_by_assignment.get((discipline_key, lesson_type, group))
        if assignment is None:
            continue

        expected_teachers = [
            teacher
            for teacher in dict.fromkeys(
                clean_cell(value)
                for value in assignment["Преподаватель нагрузки"].tolist()
            )
            if teacher
        ]
        if not expected_teachers:
            continue

        scheduled_tokens = set(
            NAME_TOKEN_PATTERN.findall(_normalize_search_text(scheduled_teacher))
        )
        expected_surnames = {
            surname
            for surname in map(_teacher_surname_key, expected_teachers)
            if surname
        }
        if scheduled_tokens.intersection(expected_surnames):
            continue

        expected_teacher_text = ", ".join(expected_teachers)
        workload_rows = {
            clean_cell(value) for value in assignment["Строка нагрузки"].tolist()
        }
        workload_rows.discard("")
        transferred_workload_rows.update(workload_rows)
        transfer_label = f"❌ По нагрузке передано: {expected_teacher_text}"
        matching_queries = [
            query
            for query, normalized_query in normalized_queries
            if normalized_query in _normalize_search_text(scheduled_teacher)
        ]
        for candidate in candidates.to_dict(orient="records"):
            for query in matching_queries:
                transferred_records.append(
                    {
                        QUERY_COLUMN: query,
                        TEACHER_COLUMN: scheduled_teacher,
                        SUGGESTED_COLUMN: transfer_label,
                        "День недели": clean_cell(candidate["День недели"]),
                        "Пара": clean_cell(candidate["Пара"]),
                        "Время": clean_cell(candidate["Время"]),
                        "Неделя": clean_cell(candidate["Неделя"]),
                        "Группа": clean_cell(candidate["Группа"]),
                        "Дисциплина": clean_cell(candidate["Дисциплина"]),
                        "Вид занятий": clean_cell(candidate["Вид занятий"]),
                        "Аудитория": clean_cell(candidate["Аудитория"]),
                        SOURCE_COLUMN: clean_cell(candidate[SOURCE_COLUMN]),
                    }
                )

        # Если в одном поиске указаны обе стороны передачи, исходная проверка
        # нагрузки уже показывает это расхождение — второй дубль не нужен.
        if any(
            normalized_query in _normalize_search_text(expected_teacher)
            for _, normalized_query in normalized_queries
            for expected_teacher in expected_teachers
        ):
            continue

        first = assignment.iloc[0]
        semesters = ", ".join(
            dict.fromkeys(clean_cell(value) for value in assignment["Семестр"].tolist())
        )
        issues.append(
            {
                "Статус": "❌ Передача нагрузки",
                "Проблема": (
                    "Занятие осталось в расписании, но по нагрузке "
                    "передано другому преподавателю"
                ),
                "Преподаватели по нагрузке": expected_teacher_text,
                "Дисциплина": clean_cell(first["Дисциплина"]),
                "Вид занятий": lesson_type,
                "Группа": group,
                "Семестр": semesters,
                "ФИО в расписании": scheduled_teacher,
                "Предлагаемое ФИО": "",
                "Возможная накладка": "",
                "Возможное место в расписании": _schedule_place(candidates),
                WORKLOAD_ROWS_KEY_COLUMN: ", ".join(sorted(workload_rows)),
            }
        )
    return issues, transferred_records, transferred_workload_rows


def audit_workload(
    workload: pd.DataFrame,
    schedule_records: pd.DataFrame,
    teacher_queries: Sequence[str],
) -> WorkloadAuditResult:
    """Сверяет только назначения выбранных преподавателей с расписанием."""

    if workload.empty:
        raise WorkloadFormatError("После фильтрации файл нагрузки пуст.")
    queries = parse_teacher_queries(teacher_queries)
    if not queries:
        raise ValueError("Введите хотя бы одну фамилию для сверки нагрузки.")
    normalized_queries = [(query, _normalize_search_text(query)) for query in queries]

    schedule = schedule_records.reindex(columns=ALL_SCHEDULE_COLUMNS).copy()
    schedule["__discipline_key"] = schedule["Дисциплина"].map(_normalize_discipline)
    schedule["__lesson_type_key"] = schedule["Вид занятий"].map(_normalize_lesson_type)
    schedule["__group_key"] = schedule["Группа"].map(
        lambda value: clean_cell(value).upper()
    )
    schedule["__teacher_surname"] = schedule[TEACHER_COLUMN].map(_teacher_surname_key)
    schedule["__day_key"] = schedule["День недели"].map(_day_sort_key)
    schedule["__pair_key"] = schedule["Пара"].map(_format_pair)
    schedule["__lesson_weeks"] = [
        _extract_lesson_weeks(discipline, week_type)
        for discipline, week_type in zip(
            schedule["Дисциплина"],
            schedule["Неделя"],
            strict=True,
        )
    ]
    empty_schedule = schedule.iloc[0:0]
    subject_candidates_by_key = {
        clean_cell(key): group
        for key, group in schedule.groupby(
            "__discipline_key",
            sort=False,
            dropna=False,
        )
    }
    exact_candidates_by_key = {
        tuple(clean_cell(value) for value in key): group
        for key, group in schedule.groupby(
            ["__discipline_key", "__lesson_type_key", "__group_key"],
            sort=False,
            dropna=False,
        )
    }

    prepared_workload = workload.copy()
    prepared_workload["__discipline_key"] = prepared_workload["Дисциплина"].map(
        _normalize_discipline
    )
    prepared_workload["__lesson_type_key"] = prepared_workload["Вид занятий"].map(
        _normalize_lesson_type
    )
    prepared_workload["__group_key"] = prepared_workload["Группа"].map(
        lambda value: clean_cell(value).upper()
    )
    assignment_columns = [
        "__discipline_key",
        "__lesson_type_key",
        "__group_key",
        "Семестр",
    ]

    error_records: list[dict[str, str]] = []
    suggestion_records: list[dict[str, str]] = []
    suggested_schedule_records: list[dict[str, str]] = []
    potential_schedule_records: list[dict[str, str]] = []
    checked_workload_rows: set[str] = set()
    failed_workload_rows: set[str] = set()

    for _, assignment in prepared_workload.groupby(
        assignment_columns,
        sort=False,
        dropna=False,
    ):
        first = assignment.iloc[0]
        discipline = clean_cell(first["Дисциплина"])
        discipline_key = clean_cell(first["__discipline_key"])
        lesson_type = _normalize_lesson_type(first["Вид занятий"])
        group = clean_cell(first["Группа"]).upper()
        semester = clean_cell(first["Семестр"])
        expected_teachers = [
            teacher
            for teacher in dict.fromkeys(
                clean_cell(value)
                for value in assignment["Преподаватель нагрузки"].tolist()
            )
            if teacher
        ]
        matching_queries = [
            query
            for query, normalized_query in normalized_queries
            if any(
                normalized_query in _normalize_search_text(teacher)
                for teacher in expected_teachers
            )
        ]
        if not matching_queries:
            continue

        relevant_teachers = [
            teacher
            for teacher in expected_teachers
            if any(
                _normalize_search_text(query) in _normalize_search_text(teacher)
                for query in matching_queries
            )
        ]
        relevant_teacher_text = ", ".join(relevant_teachers)
        relevant_surnames = {
            surname
            for surname in map(_teacher_surname_key, relevant_teachers)
            if surname
        }
        selected_assignment = assignment[
            assignment["Преподаватель нагрузки"].map(
                lambda teacher: any(
                    normalized_query in _normalize_search_text(teacher)
                    for _, normalized_query in normalized_queries
                )
            )
        ]
        workload_rows = {
            clean_cell(value)
            for value in selected_assignment["Строка нагрузки"].tolist()
        }
        workload_rows.discard("")
        checked_workload_rows.update(workload_rows)

        base_issue = {
            "Преподаватели по нагрузке": relevant_teacher_text,
            "Дисциплина": discipline,
            "Вид занятий": lesson_type,
            "Группа": group,
            "Семестр": semester,
            "ФИО в расписании": "",
            "Предлагаемое ФИО": "",
            "Возможная накладка": "",
            "Возможное место в расписании": "",
            WORKLOAD_ROWS_KEY_COLUMN: ", ".join(sorted(workload_rows)),
        }
        if not group:
            failed_workload_rows.update(workload_rows)
            error_records.append(
                {
                    "Статус": "❌ Ошибка нагрузки",
                    "Проблема": "В нагрузке не указана группа",
                    **base_issue,
                }
            )
            continue

        subject_candidates = subject_candidates_by_key.get(
            discipline_key,
            empty_schedule,
        )
        exact_candidates = exact_candidates_by_key.get(
            (discipline_key, lesson_type, group),
            empty_schedule,
        )
        if exact_candidates.empty:
            failed_workload_rows.update(workload_rows)
            problem = (
                "Дисциплина отсутствует в расписании"
                if subject_candidates.empty
                else "Не найден нужный вид занятия для группы"
            )
            error_records.append(
                {
                    "Статус": "❌ Не найдено",
                    "Проблема": problem,
                    **base_issue,
                    "Возможное место в расписании": _schedule_place(
                        subject_candidates.head(5)
                    ),
                }
            )
            continue

        for candidate in exact_candidates.to_dict(orient="records"):
            scheduled_teacher = clean_cell(candidate[TEACHER_COLUMN])
            scheduled_surname = _teacher_surname_key(scheduled_teacher)
            for relevant_teacher in relevant_teachers:
                relevant_surname = _teacher_surname_key(relevant_teacher)
                if scheduled_surname == relevant_surname:
                    workload_status = ""
                elif not scheduled_teacher and len(expected_teachers) == 1:
                    workload_status = "⚠ Однозначная подстановка из нагрузки"
                elif not scheduled_teacher:
                    workload_status = (
                        "❔ Возможное занятие: в нагрузке несколько преподавателей"
                    )
                else:
                    workload_status = (
                        "❗ В расписании указан другой преподаватель: "
                        f"{scheduled_teacher}"
                    )

                for query in matching_queries:
                    if _normalize_search_text(query) not in _normalize_search_text(
                        relevant_teacher
                    ):
                        continue
                    potential_schedule_records.append(
                        {
                            QUERY_COLUMN: query,
                            TEACHER_COLUMN: relevant_teacher,
                            SUGGESTED_COLUMN: workload_status,
                            "День недели": clean_cell(candidate["День недели"]),
                            "Пара": clean_cell(candidate["Пара"]),
                            "Время": clean_cell(candidate["Время"]),
                            "Неделя": clean_cell(candidate["Неделя"]),
                            "Группа": clean_cell(candidate["Группа"]),
                            "Дисциплина": clean_cell(candidate["Дисциплина"]),
                            "Вид занятий": clean_cell(candidate["Вид занятий"]),
                            "Аудитория": clean_cell(candidate["Аудитория"]),
                            SOURCE_COLUMN: clean_cell(candidate[SOURCE_COLUMN]),
                        }
                    )

        scheduled_teachers = [
            teacher
            for teacher in dict.fromkeys(
                clean_cell(value) for value in exact_candidates[TEACHER_COLUMN].tolist()
            )
            if teacher
        ]
        scheduled_surnames = {
            surname
            for surname in map(_teacher_surname_key, scheduled_teachers)
            if surname
        }
        if relevant_surnames and relevant_surnames.issubset(scheduled_surnames):
            continue

        blank_candidates = exact_candidates[
            exact_candidates[TEACHER_COLUMN].map(clean_cell) == ""
        ]
        if not blank_candidates.empty:
            failed_workload_rows.update(workload_rows)
            suggested_teacher = (
                expected_teachers[0] if len(expected_teachers) == 1 else ""
            )
            possible_conflict = _suggested_teacher_conflicts(
                blank_candidates,
                schedule,
                suggested_teacher,
                discipline_key,
            )
            suggestion_records.append(
                {
                    "Статус": (
                        "⚠️ Возможная накладка"
                        if possible_conflict
                        else "⚠️ Возможное место"
                    ),
                    "Проблема": (
                        "После подстановки ФИО найден другой предмет в то же время"
                        if possible_conflict
                        else "В расписании не указано ФИО"
                    ),
                    **base_issue,
                    "ФИО в расписании": ", ".join(scheduled_teachers),
                    "Предлагаемое ФИО": suggested_teacher,
                    "Возможная накладка": possible_conflict,
                    "Возможное место в расписании": _schedule_place(blank_candidates),
                }
            )
            if suggested_teacher:
                for candidate in blank_candidates.to_dict(orient="records"):
                    for query in matching_queries:
                        if _normalize_search_text(query) not in _normalize_search_text(
                            suggested_teacher
                        ):
                            continue
                        suggested_schedule_records.append(
                            {
                                QUERY_COLUMN: query,
                                TEACHER_COLUMN: suggested_teacher,
                                SUGGESTED_COLUMN: ("⚠ ФИО подставлено из нагрузки"),
                                "День недели": clean_cell(candidate["День недели"]),
                                "Пара": clean_cell(candidate["Пара"]),
                                "Время": clean_cell(candidate["Время"]),
                                "Неделя": clean_cell(candidate["Неделя"]),
                                "Группа": clean_cell(candidate["Группа"]),
                                "Дисциплина": clean_cell(candidate["Дисциплина"]),
                                "Вид занятий": clean_cell(candidate["Вид занятий"]),
                                "Аудитория": clean_cell(candidate["Аудитория"]),
                                SOURCE_COLUMN: clean_cell(candidate[SOURCE_COLUMN]),
                            }
                        )
            continue

        failed_workload_rows.update(workload_rows)
        error_records.append(
            {
                "Статус": "❌ Несовпадение ФИО",
                "Проблема": "В расписании указан другой преподаватель",
                **base_issue,
                "ФИО в расписании": ", ".join(scheduled_teachers),
                "Возможное место в расписании": _schedule_place(exact_candidates),
            }
        )

    transferred_issues, transferred_schedule_records, transferred_workload_rows = (
        _find_transferred_workload_issues(
            prepared_workload,
            schedule,
            normalized_queries,
        )
    )
    error_records.extend(transferred_issues)
    checked_workload_rows.update(transferred_workload_rows)
    failed_workload_rows.update(transferred_workload_rows)

    checked_assignments = len(checked_workload_rows)
    matched_assignments = len(checked_workload_rows - failed_workload_rows)
    errors = _collapse_workload_issue_records(error_records)
    suggestions = _collapse_workload_issue_records(suggestion_records)
    suggested_records = pd.DataFrame.from_records(
        suggested_schedule_records,
        columns=INTERNAL_COLUMNS,
    )
    if not suggested_records.empty:
        suggested_records = _deduplicate_batch_records(
            suggested_records,
            queries,
        )
    transferred_records = pd.DataFrame.from_records(
        transferred_schedule_records,
        columns=INTERNAL_COLUMNS,
    )
    if not transferred_records.empty:
        transferred_records = _deduplicate_batch_records(
            transferred_records,
            queries,
        )

    potential_records = pd.DataFrame.from_records(
        potential_schedule_records,
        columns=INTERNAL_COLUMNS,
    )
    confirmed_records = _filter_schedule_records(schedule_records, queries)
    combined_potential_records = pd.concat(
        [confirmed_records, potential_records],
        ignore_index=True,
    )
    if combined_potential_records.empty:
        potential_conflicts = pd.DataFrame(columns=DISPLAY_COLUMNS)
    else:
        combined_potential_records = combined_potential_records.drop_duplicates(
            subset=[QUERY_COLUMN, *OUTPUT_COLUMNS],
            keep="first",
            ignore_index=True,
        )
        combined_potential_records = sort_internal_schedule(
            combined_potential_records,
            queries,
        )
        potential_schedule = collapse_schedule_conflicts(combined_potential_records)
        potential_conflicts = potential_schedule[
            potential_schedule[CONFLICT_COLUMN].map(clean_cell) != ""
        ].reset_index(drop=True)

    return WorkloadAuditResult(
        checked_assignments=checked_assignments,
        matched_assignments=matched_assignments,
        errors=errors,
        suggestions=suggestions,
        suggested_records=suggested_records,
        transferred_records=transferred_records,
        potential_conflicts=potential_conflicts,
    )


def merge_workload_suggestions(
    schedule_records: pd.DataFrame,
    audit: WorkloadAuditResult | None,
    teacher_queries: Sequence[str],
) -> pd.DataFrame:
    """Добавляет к выдаче подстановки и отметки передачи нагрузки."""

    if audit is None:
        return schedule_records.reindex(columns=INTERNAL_COLUMNS)
    combined = pd.concat(
        [
            audit.transferred_records.reindex(columns=INTERNAL_COLUMNS),
            schedule_records.reindex(columns=INTERNAL_COLUMNS),
            audit.suggested_records.reindex(columns=INTERNAL_COLUMNS),
        ],
        ignore_index=True,
    )
    return _deduplicate_batch_records(combined, teacher_queries)


def merge_all_schedule_suggestions(
    schedule_records: pd.DataFrame,
    audit: WorkloadAuditResult | None,
) -> pd.DataFrame:
    """Добавляет однозначные подстановки в полный набор для проверки кабинетов."""

    original = schedule_records.reindex(columns=ALL_SCHEDULE_COLUMNS)
    if audit is None or audit.suggested_records.empty:
        return original
    combined = pd.concat(
        [
            original,
            audit.suggested_records.reindex(columns=ALL_SCHEDULE_COLUMNS),
        ],
        ignore_index=True,
    )
    return combined.drop_duplicates(
        subset=[TEACHER_COLUMN, *OUTPUT_COLUMNS],
        keep="first",
        ignore_index=True,
    )


def _pair_number(value: Any) -> int | None:
    """Возвращает положительный целочисленный номер пары для экспортной сетки."""

    pair_value = _pair_sort_key(value)
    if not math.isfinite(pair_value) or not pair_value.is_integer():
        return None
    pair_number = int(pair_value)
    return pair_number if pair_number > 0 else None


def _week_order(value: Any) -> int | None:
    """Преобразует I/II в индексы 0/1."""

    normalized_week = clean_cell(value).casefold()
    if normalized_week == "i":
        return 0
    if normalized_week == "ii":
        return 1
    return None


def _teacher_display_name(records: pd.DataFrame, query: str) -> str:
    """Подставляет полное ФИО, если запрос однозначно ему соответствует."""

    query_records = records[
        records[QUERY_COLUMN].map(_normalize_search_text)
        == _normalize_search_text(query)
    ]
    candidates: list[str] = []
    for teacher_value in query_records[TEACHER_COLUMN].tolist():
        teacher = clean_cell(teacher_value)
        parts = [
            clean_cell(part)
            for part in TEACHER_NAME_SEPARATOR_PATTERN.split(teacher)
            if clean_cell(part)
        ]
        matching_parts = [
            part
            for part in parts
            if _normalize_search_text(query) in _normalize_search_text(part)
        ]
        candidates.extend(matching_parts or ([teacher] if teacher else []))

    if not candidates:
        return clean_cell(query)

    surname_keys = {
        _normalize_search_text(candidate.split(maxsplit=1)[0])
        for candidate in candidates
        if candidate.split(maxsplit=1)
    }
    if len(surname_keys) > 1:
        return clean_cell(query)

    counts = Counter(candidates)
    return max(counts, key=lambda candidate: (counts[candidate], len(candidate)))


def _export_clusters_for_query(
    records: pd.DataFrame,
    query: str,
) -> list[dict[str, Any]]:
    """Собирает группы совместного занятия, не смешивая разные занятия слота."""

    query_records = records[
        records[QUERY_COLUMN].map(_normalize_search_text)
        == _normalize_search_text(query)
    ]
    clusters: dict[tuple[Any, ...], dict[str, Any]] = {}

    for record in query_records.to_dict(orient="records"):
        day_order = _day_sort_key(record["День недели"])
        pair_number = _pair_number(record["Пара"])
        week_order = _week_order(record["Неделя"])
        if day_order > 5 or pair_number is None or week_order is None:
            continue

        lesson_type = clean_cell(record["Вид занятий"])
        discipline = clean_cell(record["Дисциплина"])
        group = clean_cell(record["Группа"])
        non_lecture_group_key = (
            "" if _normalize_lesson_type(lesson_type) == "ЛК" else group.upper()
        )

        key = (
            day_order,
            pair_number,
            week_order,
            _teacher_conflict_identity(query, record[TEACHER_COLUMN]),
            _normalize_search_text(record["Время"]),
            _normalize_room_conflict_discipline(discipline),
            _normalize_search_text(record["Вид занятий"]),
            _normalize_search_text(record["Аудитория"]),
            non_lecture_group_key,
        )
        cluster = clusters.setdefault(
            key,
            {
                "day_order": day_order,
                "pair_number": pair_number,
                "week_order": week_order,
                "teacher_identity": key[3],
                "teacher": clean_cell(record[TEACHER_COLUMN]),
                "suggested": False,
                "workload_error": False,
                "time": clean_cell(record["Время"]),
                "discipline": discipline,
                "lesson_type": lesson_type,
                "auditorium": clean_cell(record["Аудитория"]),
                "groups": [],
                "weeks": set(),
            },
        )
        cluster["weeks"].update(_extract_lesson_weeks(discipline, record["Неделя"]))
        workload_status = clean_cell(record.get(SUGGESTED_COLUMN, ""))
        is_workload_error = workload_status.startswith("❌")
        cluster["workload_error"] = cluster["workload_error"] or is_workload_error
        cluster["suggested"] = cluster["suggested"] or (
            bool(workload_status) and not is_workload_error
        )
        if group and group not in cluster["groups"]:
            cluster["groups"].append(group)

    result = list(clusters.values())
    result.sort(
        key=lambda cluster: (
            cluster["day_order"],
            cluster["pair_number"],
            cluster["week_order"],
            cluster["discipline"].casefold(),
            ", ".join(cluster["groups"]).casefold(),
            cluster["auditorium"].casefold(),
        )
    )
    return result


def _append_room_conflicts_sheet(
    workbook: Workbook,
    room_conflicts: pd.DataFrame,
) -> None:
    """Добавляет в книгу отдельный оформленный лист конфликтов кабинетов."""

    if room_conflicts.empty:
        return

    worksheet = workbook.create_sheet("Накладки аудиторий")
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 80
    worksheet.freeze_panes = "A4"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    total_columns = len(ROOM_CONFLICT_COLUMNS)
    last_column_letter = get_column_letter(total_columns)
    title_fill = PatternFill("solid", fgColor="9C0006")
    header_fill = PatternFill("solid", fgColor="C00000")
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    conflict_fill = PatternFill("solid", fgColor="FCE4D6")
    thin_red = Side(style="thin", color="C00000")
    border = Border(left=thin_red, right=thin_red, top=thin_red, bottom=thin_red)

    worksheet.merge_cells(f"A1:{last_column_letter}1")
    worksheet["A1"] = "Накладки аудиторий выбранных преподавателей"
    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    worksheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.merge_cells(f"A2:{last_column_letter}2")
    worksheet["A2"] = (
        "Проверка выполнена по всем загруженным файлам. Общая лекция одного "
        "предмета для нескольких групп не считается конфликтом."
    )
    worksheet["A2"].fill = note_fill
    worksheet["A2"].font = Font(name="Arial", size=10, italic=True, color="7F6000")
    worksheet["A2"].alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    for column_index, column_name in enumerate(ROOM_CONFLICT_COLUMNS, start=1):
        cell = worksheet.cell(3, column_index, column_name)
        cell.fill = header_fill
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = border

    export_frame = room_conflicts.reindex(columns=ROOM_CONFLICT_COLUMNS).fillna("")
    for row_index, record in enumerate(
        export_frame.itertuples(index=False, name=None),
        start=4,
    ):
        maximum_lines = 1
        for column_index, value in enumerate(record, start=1):
            text_value = _excel_safe_text(value, multiline=True)
            maximum_lines = max(maximum_lines, len(text_value.splitlines()))
            cell = worksheet.cell(row_index, column_index, text_value)
            cell.fill = conflict_fill
            cell.font = Font(
                name="Arial",
                size=10,
                bold=column_index in {7, 8},
                color="9C0006" if column_index == 8 else "000000",
            )
            cell.alignment = Alignment(
                horizontal="center" if column_index in {2, 3, 4, 5, 6, 7} else "left",
                vertical="top",
                wrap_text=True,
            )
            cell.border = border
        worksheet.row_dimensions[row_index].height = min(
            360,
            max(30, 18 + maximum_lines * 18),
        )

    worksheet.row_dimensions[1].height = 30
    worksheet.row_dimensions[2].height = 34
    worksheet.row_dimensions[3].height = 36
    column_widths = (26, 14, 8, 18, 10, 18, 22, 28, 32, 26, 38, 16, 42)
    for column_index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.auto_filter.ref = f"A3:{last_column_letter}{worksheet.max_row}"
    worksheet.print_title_rows = "1:3"
    worksheet.print_area = f"A1:{last_column_letter}{worksheet.max_row}"


def build_schedule_xlsx(
    records: pd.DataFrame,
    teacher_queries: Sequence[str],
    room_conflicts: pd.DataFrame | None = None,
) -> bytes:
    """Создаёт цветную матрицу и лист конфликтов кабинетов в XLSX."""

    queries = parse_teacher_queries(teacher_queries)
    if not queries:
        raise ScheduleExportError("Не выбраны преподаватели для выгрузки.")
    if records.empty:
        raise ScheduleExportError("Нет найденных занятий для выгрузки.")

    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Расписание"
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_view.zoomScale = 80
        worksheet.freeze_panes = "D3"
        worksheet.print_title_rows = "1:2"
        worksheet.print_title_cols = "A:C"
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        workbook.properties.title = "Расписание преподавателей"
        workbook.properties.creator = "Streamlit Schedule Parser"

        clusters_by_query = {
            _normalize_search_text(query): _export_clusters_for_query(records, query)
            for query in queries
        }
        maximum_pair = max(
            7,
            max(
                (
                    cluster["pair_number"]
                    for clusters in clusters_by_query.values()
                    for cluster in clusters
                ),
                default=7,
            ),
        )
        total_columns = 3 + len(queries) * 4
        rows_per_day = maximum_pair * 2
        total_rows = 2 + len(DAY_EXPORT_CONFIG) * rows_per_day

        thin_gray = Side(style="thin", color="A6A6A6")
        medium_black = Side(style="medium", color="000000")
        medium_red = Side(style="medium", color="C00000")
        header_fill = PatternFill("solid", fgColor="D3DFE2")
        teacher_fill = PatternFill("solid", fgColor="75FB4C")
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        suggested_fill = PatternFill("solid", fgColor="FFF2CC")
        conflict_fill = PatternFill("solid", fgColor="F4CCCC")

        teacher_starts = {
            4 + teacher_index * 4 for teacher_index in range(len(queries))
        }
        teacher_ends = {7 + teacher_index * 4 for teacher_index in range(len(queries))}
        medium_left_columns = {1, *teacher_starts}
        medium_right_columns = {3, *teacher_ends, total_columns}
        day_start_rows = {
            3 + day_index * rows_per_day for day_index in range(len(DAY_EXPORT_CONFIG))
        }
        day_end_rows = {
            2 + (day_index + 1) * rows_per_day
            for day_index in range(len(DAY_EXPORT_CONFIG))
        }

        def cell_border(
            row: int,
            column: int,
            *,
            conflict: bool = False,
        ) -> Border:
            if conflict:
                return Border(
                    left=medium_red,
                    right=medium_red,
                    top=medium_red,
                    bottom=medium_red,
                )
            return Border(
                left=medium_black if column in medium_left_columns else thin_gray,
                right=medium_black if column in medium_right_columns else thin_gray,
                top=(medium_black if row == 1 or row in day_start_rows else thin_gray),
                bottom=(
                    medium_black
                    if row == 2 or row in day_end_rows or row == total_rows
                    else thin_gray
                ),
            )

        for column, value in enumerate(("день", "№", "Нед"), start=1):
            worksheet.merge_cells(
                start_row=1,
                start_column=column,
                end_row=2,
                end_column=column,
            )
            worksheet.cell(1, column, value)
            for row in (1, 2):
                cell = worksheet.cell(row, column)
                cell.fill = header_fill
                cell.font = Font(name="Arial", size=12, bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

        for teacher_index, query in enumerate(queries):
            start_column = 4 + teacher_index * 4
            end_column = start_column + 3
            clusters = clusters_by_query[_normalize_search_text(query)]
            display_name = _teacher_display_name(records, query)
            header_value = f"{display_name}\n(занятий: {len(clusters)})"
            worksheet.merge_cells(
                start_row=1,
                start_column=start_column,
                end_row=1,
                end_column=end_column,
            )
            worksheet.cell(
                1, start_column, _excel_safe_text(header_value, multiline=True)
            )

            for column in range(start_column, end_column + 1):
                header_cell = worksheet.cell(1, column)
                header_cell.fill = teacher_fill
                header_cell.font = Font(name="Arial", size=14, bold=True)
                header_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

            for offset, subheader in enumerate(EXPORT_SUBHEADERS):
                subheader_cell = worksheet.cell(2, start_column + offset, subheader)
                subheader_cell.fill = white_fill
                subheader_cell.font = Font(name="Arial", size=11, bold=False)
                subheader_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

        slot_clusters: dict[tuple[str, int, int, int], list[dict[str, Any]]] = {}
        for query in queries:
            normalized_query = _normalize_search_text(query)
            for cluster in clusters_by_query[normalized_query]:
                slot_key = (
                    normalized_query,
                    cluster["day_order"],
                    cluster["pair_number"],
                    cluster["week_order"],
                )
                slot_clusters.setdefault(slot_key, []).append(cluster)

        for day_index, (day_label, day_color, odd_color, even_color) in enumerate(
            DAY_EXPORT_CONFIG
        ):
            day_start_row = 3 + day_index * rows_per_day
            day_end_row = day_start_row + rows_per_day - 1
            worksheet.cell(day_start_row, 1, day_label)

            for pair_number in range(1, maximum_pair + 1):
                pair_start_row = day_start_row + (pair_number - 1) * 2
                worksheet.cell(pair_start_row, 2, pair_number)

                for week_order in (0, 1):
                    row = pair_start_row + week_order
                    worksheet.cell(row, 3, week_order + 1)
                    body_color = odd_color if pair_number % 2 else even_color
                    body_fill = PatternFill("solid", fgColor=body_color)
                    day_fill = PatternFill("solid", fgColor=day_color)
                    maximum_lines = 0

                    for column in range(1, total_columns + 1):
                        cell = worksheet.cell(row, column)
                        cell.fill = day_fill if column == 1 else body_fill
                        cell.font = Font(name="Arial", size=10)
                        cell.alignment = Alignment(
                            horizontal="center" if column <= 3 else "left",
                            vertical="center",
                            wrap_text=True,
                        )

                    for teacher_index, query in enumerate(queries):
                        start_column = 4 + teacher_index * 4
                        clusters = slot_clusters.get(
                            (
                                _normalize_search_text(query),
                                day_index,
                                pair_number,
                                week_order,
                            ),
                            [],
                        )
                        lesson_counts_by_teacher: dict[str, Counter[int]] = {}
                        for cluster in clusters:
                            week_counts = lesson_counts_by_teacher.setdefault(
                                cluster["teacher_identity"],
                                Counter(),
                            )
                            week_counts.update(cluster["weeks"])
                        conflicting_teachers = {
                            teacher_identity
                            for teacher_identity, week_counts in (
                                lesson_counts_by_teacher.items()
                            )
                            if max(week_counts.values(), default=0) > 1
                        }
                        has_overlap = bool(conflicting_teachers)
                        has_workload_error = any(
                            cluster["workload_error"] for cluster in clusters
                        )
                        has_suggestion = any(
                            cluster["suggested"] for cluster in clusters
                        )
                        maximum_lines = max(maximum_lines, len(clusters))
                        disciplines = [cluster["discipline"] for cluster in clusters]
                        if has_overlap and disciplines:
                            warning_index = next(
                                index
                                for index, cluster in enumerate(clusters)
                                if cluster["teacher_identity"] in conflicting_teachers
                            )
                            disciplines[warning_index] = (
                                f"⚠ {disciplines[warning_index]}"
                            )
                        elif has_workload_error and disciplines:
                            warning_index = next(
                                index
                                for index, cluster in enumerate(clusters)
                                if cluster["workload_error"]
                            )
                            disciplines[warning_index] = (
                                f"❌ {disciplines[warning_index]}"
                            )
                        elif has_suggestion and disciplines:
                            warning_index = next(
                                index
                                for index, cluster in enumerate(clusters)
                                if cluster["suggested"]
                            )
                            disciplines[warning_index] = (
                                f"⚠ {disciplines[warning_index]}"
                            )
                        values = (
                            "\n".join(disciplines),
                            "\n".join(cluster["lesson_type"] for cluster in clusters),
                            "\n".join(
                                ", ".join(cluster["groups"]) for cluster in clusters
                            ),
                            "\n".join(cluster["auditorium"] for cluster in clusters),
                        )
                        for offset, value in enumerate(values):
                            data_cell = worksheet.cell(
                                row,
                                start_column + offset,
                                _excel_safe_text(value, multiline=True),
                            )
                            data_cell.alignment = Alignment(
                                horizontal="left" if offset == 0 else "center",
                                vertical="center",
                                wrap_text=True,
                            )
                            data_cell.border = cell_border(
                                row,
                                start_column + offset,
                                conflict=has_overlap or has_workload_error,
                            )
                            if has_overlap or has_workload_error:
                                data_cell.fill = conflict_fill
                            elif has_suggestion:
                                data_cell.fill = suggested_fill
                            if (has_overlap or has_workload_error) and offset == 0:
                                data_cell.font = Font(
                                    name="Arial",
                                    size=10,
                                    bold=True,
                                    color="C00000",
                                )
                            elif has_suggestion and offset == 0:
                                data_cell.font = Font(
                                    name="Arial",
                                    size=10,
                                    bold=True,
                                    color="9C6500",
                                )

                    if maximum_lines:
                        worksheet.row_dimensions[row].height = min(
                            400,
                            38 + (maximum_lines - 1) * 24,
                        )
                    else:
                        worksheet.row_dimensions[row].height = 24

            for row in range(day_start_row, day_end_row + 1):
                worksheet.cell(row, 1).fill = PatternFill("solid", fgColor=day_color)
            worksheet.merge_cells(
                start_row=day_start_row,
                start_column=1,
                end_row=day_end_row,
                end_column=1,
            )
            worksheet.cell(day_start_row, 1).font = Font(
                name="Arial",
                size=12,
                bold=True,
            )
            worksheet.cell(day_start_row, 1).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            for pair_number in range(1, maximum_pair + 1):
                pair_start_row = day_start_row + (pair_number - 1) * 2
                worksheet.merge_cells(
                    start_row=pair_start_row,
                    start_column=2,
                    end_row=pair_start_row + 1,
                    end_column=2,
                )
                worksheet.cell(pair_start_row, 2).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

        for row in range(1, total_rows + 1):
            for column in range(1, total_columns + 1):
                cell = worksheet.cell(row, column)
                if cell.border == Border():
                    cell.border = cell_border(row, column)

        worksheet.row_dimensions[1].height = 72
        worksheet.row_dimensions[2].height = 26
        worksheet.column_dimensions["A"].width = 8
        worksheet.column_dimensions["B"].width = 5
        worksheet.column_dimensions["C"].width = 6
        teacher_widths = (32, 9, 16, 20)
        for teacher_index in range(len(queries)):
            start_column = 4 + teacher_index * 4
            for offset, width in enumerate(teacher_widths):
                worksheet.column_dimensions[
                    get_column_letter(start_column + offset)
                ].width = width

        worksheet.print_area = f"A1:{get_column_letter(total_columns)}{total_rows}"

        if room_conflicts is not None:
            _append_room_conflicts_sheet(workbook, room_conflicts)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    except ScheduleExportError:
        raise
    except Exception as exc:
        LOGGER.exception("Не удалось сформировать XLSX")
        raise ScheduleExportError(
            "Не удалось сформировать Excel-файл с расписанием."
        ) from exc


def _render_workload_audit(
    st: Any,
    audit: WorkloadAuditResult,
    academic_term: str,
    teacher_queries: Sequence[str],
) -> None:
    """Показывает сворачиваемую сверку только выбранных преподавателей."""

    st.subheader("Сверка с учебной нагрузкой")
    parity_text = (
        "нечётные" if _normalize_search_text(academic_term) == "осень" else "чётные"
    )
    teacher_text = ", ".join(parse_teacher_queries(teacher_queries))
    st.caption(
        f"Проверены только преподаватели: {teacher_text}; {parity_text} семестры. "
        "В XLSX добавляются только однозначные подстановки ФИО в уже существующие "
        "занятия расписания."
    )
    if audit.checked_assignments == 0:
        st.info("В нагрузке не найдены назначения выбранных преподавателей.")
        return

    summary_columns = st.columns(5)
    summary_columns[0].metric("Проверено", audit.checked_assignments)
    summary_columns[1].metric("Совпало", audit.matched_assignments)
    summary_columns[2].metric("Ошибок", len(audit.errors.index))
    summary_columns[3].metric("Возможных мест", len(audit.suggestions.index))
    summary_columns[4].metric("Накладок", len(audit.potential_conflicts.index))

    if (
        audit.errors.empty
        and audit.suggestions.empty
        and audit.potential_conflicts.empty
    ):
        st.success("Нагрузка соответствует загруженному расписанию.")
        return
    if not audit.errors.empty:
        with st.expander(
            f"❌ Ошибки сверки нагрузки ({len(audit.errors.index)})",
            expanded=False,
        ):
            st.caption("Показаны замечания только по введённым преподавателям.")
            st.dataframe(
                audit.errors,
                width="stretch",
                hide_index=True,
                row_height=54,
            )
    if not audit.suggestions.empty:
        with st.expander(
            f"⚠ Возможные места и подстановки ({len(audit.suggestions.index)})",
            expanded=False,
        ):
            st.caption(
                "Однозначная подстановка попадает в XLSX; неоднозначная "
                "остаётся только подсказкой на сайте."
            )
            st.dataframe(
                audit.suggestions,
                width="stretch",
                hide_index=True,
                row_height=70,
            )

        exported_count = len(audit.suggested_records.index)
        if exported_count:
            st.warning(
                f"В сводное расписание добавлено однозначно восстановленных "
                f"строк: {exported_count}. Они отмечены знаком ⚠ и цветом."
            )

    if not audit.potential_conflicts.empty:
        maximum_lessons = max(
            (
                len(_clean_multiline_cell(value).splitlines())
                for value in audit.potential_conflicts["Дисциплина"]
            ),
            default=1,
        )
        with st.expander(
            f"🔴 Возможные накладки по расписанию и нагрузке "
            f"({len(audit.potential_conflicts.index)})",
            expanded=False,
        ):
            st.caption(
                "Здесь учитываются все возможные места выбранного преподавателя. "
                "Неоднозначные варианты показаны только для проверки и не "
                "добавляются в XLSX."
            )
            st.dataframe(
                audit.potential_conflicts.reindex(columns=DISPLAY_COLUMNS),
                width="stretch",
                hide_index=True,
                row_height=min(400, max(70, 30 + maximum_lessons * 24)),
            )


def _style_schedule_table(frame: pd.DataFrame) -> Any:
    """Выделяет накладки и замечания нагрузки в расписании сайта."""

    def row_style(row: pd.Series) -> list[str]:
        workload_status = clean_cell(row.get(SUGGESTED_COLUMN, ""))
        conflict_status = clean_cell(row.get(CONFLICT_COLUMN, ""))
        if conflict_status or workload_status.startswith("❌"):
            css = "background-color: #FCE8E6; color: #9C0006;"
        elif workload_status:
            css = "background-color: #FFF4CE; color: #7A5200;"
        else:
            css = ""
        return [css] * len(row.index)

    return frame.style.apply(row_style, axis=1)


def _render_footer(st: Any) -> None:
    """Показывает постоянный нижний блок со справкой и авторством."""

    st.divider()
    help_column, version_column, author_column = st.columns([1.2, 1.6, 7.2])
    with help_column:
        with st.popover("❓ Справка"):
            st.markdown("### Как пользоваться")
            st.markdown(
                """
1. Нажмите **Browse files** и выберите один или несколько файлов `.xls`/`.xlsx`.
2. При необходимости загрузите отдельный файл нагрузки и выберите **Осень** или **Весна**.
3. Введите фамилии преподавателей — каждую с новой строки либо через запятую.
4. Нажмите **Найти расписание**.
5. Раскрывайте только нужные разделы: расписание, ошибки нагрузки, возможные подстановки или накладки.
6. Жёлтый значок **⚠** означает, что ФИО однозначно восстановлено по нагрузке. Только такие подстановки добавляются в XLSX; неоднозначные варианты остаются на сайте.
7. Сверка нагрузки выполняется в обе стороны: если занятие осталось в расписании у преподавателя, но в нагрузке уже передано другому, оно получает красную отметку на сайте и в XLSX.
8. Красным цветом отмечаются два и более занятия преподавателя в одно время и на пересекающихся учебных неделях. Практики разных групп считаются отдельно, а общая лекция нескольких групп остаётся одним занятием.
9. Блок **Накладки аудиторий** показывает, если один кабинет одновременно занят разными занятиями. Общая лекция одного предмета для нескольких групп исключается из предупреждений.
10. Нажмите **Скачать сводное расписание XLSX**, чтобы получить цветную сетку и отдельный лист с найденными накладками аудиторий.

**Как это работает:** приложение написано на Python и Streamlit. Pandas читает загруженные Excel-файлы, очищает объединённые и «грязные» ячейки, после чего поиск собирает занятия выбранных преподавателей из всех курсов.
                """
            )
    with version_column:
        with st.popover(f"Версия {APP_VERSION}"):
            st.markdown("### История изменений")
            for version, release_date, changes in APP_CHANGELOG:
                st.markdown(f"**{version} — {release_date}**")
                for change in changes:
                    st.markdown(f"- {change}")
    with author_column:
        st.caption("Автор: Трушин С.")


def _render_app(st: Any) -> None:
    """Отрисовывает основную часть пользовательского интерфейса."""

    st.title("Расписание преподавателя")
    st.caption(
        "Загрузите расписания нескольких курсов, укажите нужных преподавателей "
        "и скачайте сводную цветную сетку."
    )

    uploaded_files = st.file_uploader(
        "Excel-файлы расписания",
        type=["xls", "xlsx"],
        accept_multiple_files=True,
        help=(
            "Можно выбрать сразу несколько файлов .xls/.xlsx; "
            "в каждом обрабатывается первый лист."
        ),
        key="schedule_files",
    )
    academic_term = st.radio(
        "Текущий учебный семестр",
        options=("Осень", "Весна"),
        horizontal=True,
        help=("Для осени проверяются нечётные семестры нагрузки, для весны — чётные."),
    )
    workload_file = st.file_uploader(
        "Файл учебной нагрузки — необязательно",
        type=["xls", "xlsx"],
        accept_multiple_files=False,
        help=("Если файл не загружен, приложение работает только с расписанием."),
        key="workload_file",
    )
    teacher_input = st.text_area(
        "Преподаватели",
        placeholder="Иванов\nПетров\nСидорова",
        help=(
            "Введите фамилии или части ФИО с новой строки, через запятую "
            "или точку с запятой."
        ),
        height=120,
    )
    button_clicked = st.button("Найти расписание", type="primary")
    search_requested = button_clicked or bool(teacher_input.strip())

    if not search_requested:
        return
    if not uploaded_files:
        st.info("Сначала загрузите хотя бы один Excel-файл с расписанием.")
        return
    teacher_queries = parse_teacher_queries(teacher_input)
    if not teacher_queries:
        st.warning("Введите хотя бы одну фамилию преподавателя.")
        return

    with st.spinner("Читаю файл и ищу занятия…"):
        try:
            batch_result = parse_schedule_files(
                [
                    (uploaded_file.name, uploaded_file.getvalue())
                    for uploaded_file in uploaded_files
                ],
                teacher_queries,
                st.session_state,
            )
        except (ScheduleError, ValueError) as exc:
            st.error(str(exc))
            return
        except Exception:
            LOGGER.exception("Непредвиденная ошибка обработки расписания")
            st.error(
                "Не удалось обработать расписание из-за непредвиденной ошибки. "
                "Проверьте структуру файла и попробуйте снова."
            )
            return

    workload_audit: WorkloadAuditResult | None = None
    workload_error = ""
    if workload_file is not None:
        with st.spinner("Сверяю учебную нагрузку с расписанием…"):
            try:
                workload = parse_workload_file(
                    workload_file.getvalue(),
                    academic_term,
                )
                workload_audit = audit_workload(
                    workload,
                    batch_result.all_records,
                    teacher_queries,
                )
            except (ScheduleError, ValueError) as exc:
                workload_error = str(exc)
            except Exception:
                LOGGER.exception("Непредвиденная ошибка сверки нагрузки")
                workload_error = (
                    "Не удалось сверить нагрузку из-за непредвиденной ошибки."
                )

    if batch_result.errors:
        st.warning("Некоторые файлы не обработаны. Остальные результаты сохранены.")
        with st.expander("Ошибки файлов"):
            for file_error in batch_result.errors:
                st.write(f"• {file_error.filename}: {file_error.message}")

    if batch_result.duplicate_files:
        duplicate_names = ", ".join(batch_result.duplicate_files)
        st.info(f"Повторно загруженные файлы пропущены: {duplicate_names}")

    if not batch_result.processed_files:
        st.error("Не удалось обработать ни одного загруженного файла.")
        return

    if workload_error:
        st.error(f"Ошибка файла нагрузки: {workload_error}")
    elif workload_audit is not None:
        _render_workload_audit(
            st,
            workload_audit,
            academic_term,
            teacher_queries,
        )

    result_df = merge_workload_suggestions(
        batch_result.records,
        workload_audit,
        teacher_queries,
    )
    room_schedule_records = merge_all_schedule_suggestions(
        batch_result.all_records,
        workload_audit,
    )
    room_conflicts = find_room_conflicts(
        room_schedule_records,
        teacher_queries,
    )
    if result_df.empty:
        st.warning("Занятия указанных преподавателей не найдены.")
        return

    found_queries = {
        _normalize_search_text(value) for value in result_df[QUERY_COLUMN].tolist()
    }
    missing_queries = [
        query
        for query in teacher_queries
        if _normalize_search_text(query) not in found_queries
    ]
    st.success(
        f"Найдено записей: {len(result_df)}; "
        f"обработано файлов: {len(batch_result.processed_files)}."
    )
    if missing_queries:
        st.info("Без совпадений: " + ", ".join(missing_queries))

    display_df = collapse_schedule_conflicts(result_df)
    conflict_count = int(display_df[CONFLICT_COLUMN].astype(bool).sum())
    maximum_lessons_per_slot = max(
        (
            len(_clean_multiline_cell(value).splitlines())
            for value in display_df["Дисциплина"]
        ),
        default=1,
    )
    schedule_row_height = min(
        400,
        max(54, 30 + maximum_lessons_per_slot * 24),
    )
    if conflict_count:
        st.warning(
            f"Обнаружено накладок: {conflict_count}. Разные занятия одного "
            "временного слота показаны вместе в одной строке."
        )

    with st.expander(
        f"Расписание ({len(display_df.index)} строк)",
        expanded=True,
    ):
        st.dataframe(
            _style_schedule_table(display_df.reindex(columns=DISPLAY_COLUMNS)),
            width="stretch",
            hide_index=True,
            row_height=schedule_row_height,
        )

    if conflict_count:
        with st.expander(
            f"🔴 Накладки выбранных преподавателей ({conflict_count})",
            expanded=False,
        ):
            st.dataframe(
                _style_schedule_table(
                    display_df[display_df[CONFLICT_COLUMN].astype(bool)].reindex(
                        columns=DISPLAY_COLUMNS
                    )
                ),
                width="stretch",
                hide_index=True,
                row_height=schedule_row_height,
            )

    if room_conflicts.empty:
        st.success("Накладок аудиторий для выбранных преподавателей не найдено.")
    else:
        maximum_room_lessons = max(
            (
                len(_clean_multiline_cell(value).splitlines())
                for value in room_conflicts["Дисциплины"]
            ),
            default=2,
        )
        st.warning(
            f"Обнаружено накладок аудиторий: {len(room_conflicts.index)}. "
            "Проверка учитывает все загруженные расписания."
        )
        with st.expander(
            f"🏫 Накладки аудиторий ({len(room_conflicts.index)})",
            expanded=False,
        ):
            st.caption(
                "Показаны только конфликты, затрагивающие введённых "
                "преподавателей. Общая лекция одного предмета для нескольких "
                "групп не считается ошибкой."
            )
            st.dataframe(
                room_conflicts.reindex(columns=ROOM_CONFLICT_COLUMNS),
                width="stretch",
                hide_index=True,
                row_height=min(400, max(70, 30 + maximum_room_lessons * 24)),
            )

    try:
        export_bytes = build_schedule_xlsx(
            result_df,
            teacher_queries,
            room_conflicts,
        )
    except ScheduleError as exc:
        st.error(str(exc))
        return

    st.download_button(
        "Скачать сводное расписание XLSX",
        data=export_bytes,
        file_name="расписание_преподавателей.xlsx",
        mime=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        type="primary",
    )


def main() -> None:
    """Запускает постоянно доступный пользовательский интерфейс Streamlit."""

    try:
        import streamlit as st
    except ModuleNotFoundError as exc:  # pragma: no cover - подсказка для CLI-запуска
        raise RuntimeError(
            "Streamlit не установлен. Выполните 'pip install -r requirements.txt', "
            "затем запустите 'streamlit run app.py'."
        ) from exc

    st.set_page_config(
        page_title="Поиск расписания преподавателя",
        layout="wide",
    )
    try:
        _render_app(st)
    finally:
        _render_footer(st)


if __name__ == "__main__":
    main()
